
"""
Dynamic Ranch Inventory Service
===============================

Comprehensive service layer for the Ranch Management Inventory Engine.

Place at:
backend/app/services/dynamic_inventory_service.py

Main rule:
Users enter INPUT values only. The system calculates OUTPUT values automatically
from department templates and system-owned calculation rules.

Departments:
- Crops Department
- Animals Department
- Machineries & Maintenance
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import re
from calendar import monthrange
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple, Union

from sqlalchemy import func, or_
from sqlalchemy.orm import Session, joinedload

try:
    from fastapi import HTTPException
    from fastapi.responses import StreamingResponse
except Exception:  # pragma: no cover
    HTTPException = Exception  # type: ignore
    StreamingResponse = None  # type: ignore

try:
    from passlib.context import CryptContext
except Exception:  # pragma: no cover
    CryptContext = None  # type: ignore

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    Workbook = None  # type: ignore

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
except Exception:  # pragma: no cover
    colors = None  # type: ignore

from app.models.dynamic_inventory import (
    RANCH_CLASSIFICATION_FIELDS,
    DynamicInventory,
    DynamicInventoryAlert,
    DynamicInventoryAttachment,
    DynamicInventoryAuditLog,
    DynamicInventoryCalculationRule,
    DynamicInventoryCredential,
    DynamicInventoryField,
    DynamicInventoryLookupOption,
    DynamicInventoryMetric,
    DynamicInventoryPeriod,
    DynamicInventoryReport,
    DynamicInventoryRow,
    DynamicInventoryTemplate,
    DynamicInventoryTemplateCalculationRule,
    DynamicInventoryTemplateField,
    DynamicInventoryTemplateMetric,
    DynamicInventoryUserAccess,
    DynamicInventoryValue,
    InventoryAccessType,
    InventoryAlertLevel,
    InventoryAlertStatus,
    InventoryAuditAction,
    InventoryCalculationRuleType,
    InventoryCalculationScope,
    InventoryFieldCategory,
    InventoryFieldDirection,
    InventoryFieldType,
    InventoryLookupGroup,
    InventoryMetricType,
    InventoryPeriodStatus,
    InventoryPeriodType,
    InventoryReportFormat,
    InventoryReportStatus,
    InventoryReportType,
    InventoryStatus,
    InventoryTemplateType,
    InventoryUserRole,
    RanchDepartment,
)


# =============================================================================
# Generic helpers
# =============================================================================


def _now() -> datetime:
    return datetime.now(timezone.utc)


def _dump(payload: Any, *, exclude_unset: bool = False) -> Dict[str, Any]:
    if payload is None:
        return {}
    if isinstance(payload, dict):
        return dict(payload)
    if hasattr(payload, "model_dump"):
        return payload.model_dump(exclude_unset=exclude_unset)
    if hasattr(payload, "dict"):
        return payload.dict(exclude_unset=exclude_unset)
    return dict(payload)


def _to_enum(enum_cls: Any, value: Any, default: Any = None) -> Any:
    if value is None:
        return default
    if isinstance(value, enum_cls):
        return value
    raw = str(value).strip()
    for item in enum_cls:
        if raw == item.name or raw.upper() == item.name or raw.lower() == str(item.value).lower():
            return item
    if default is not None:
        return default
    raise ValueError(f"Invalid {enum_cls.__name__}: {value}")


def _enum_value(value: Any) -> str:
    """Return a safe string value for Enum/string calculation keys."""
    if value is None:
        return ""
    if hasattr(value, "value"):
        return str(value.value)
    return str(value)


def _calc_type(name: str) -> Any:
    """
    Resolve a calculation enum member safely without recursion.

    Some deployed model files may not yet include every ranch calculation enum
    member. In that case, seed the rule as CUSTOM_SYSTEM_RULE while preserving
    the rule_key; the calculation engine still uses rule_key to run the correct
    system-owned calculation.
    """
    member = getattr(InventoryCalculationRuleType, name, None)
    if member is not None:
        return member

    fallback = getattr(InventoryCalculationRuleType, "CUSTOM_SYSTEM_RULE", None)
    if fallback is not None:
        return fallback

    # Last-resort fallback for older model files that do not yet define
    # CUSTOM_SYSTEM_RULE. This avoids startup/seed crashes; the rule_key still
    # preserves the real calculation name.
    try:
        return next(iter(InventoryCalculationRuleType))
    except StopIteration:
        return name


def _slugify(value: str) -> str:
    value = (value or "").strip().lower()
    value = re.sub(r"[^a-z0-9]+", "-", value)
    value = re.sub(r"-+", "-", value).strip("-")
    return value or "inventory"


def _normalize_key(value: str) -> str:
    value = (value or "").strip().lower()
    value = re.sub(r"[^a-z0-9]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    return value or "field"


def _field_code(index: int) -> str:
    index += 1
    code = ""
    while index:
        index, rem = divmod(index - 1, 26)
        code = chr(65 + rem) + code
    return code


def _decimal_or_none(value: Any) -> Optional[Decimal]:
    if value is None or value == "":
        return None
    try:
        if isinstance(value, bool):
            return Decimal(1 if value else 0)
        return Decimal(str(value).replace(",", ""))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _num(value: Any, default: Decimal = Decimal("0")) -> Decimal:
    parsed = _decimal_or_none(value)
    return parsed if parsed is not None else default


def _safe_divide(a: Decimal, b: Decimal) -> Optional[Decimal]:
    if b is None or b == 0:
        return None
    return a / b


def _format_decimal(value: Optional[Decimal]) -> Optional[str]:
    if value is None:
        return None
    if value == value.to_integral():
        return str(value.quantize(Decimal("1")))
    return str(value.normalize())


def _jsonable(value: Any) -> Any:
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, list):
        return [_jsonable(v) for v in value]
    if isinstance(value, dict):
        return {k: _jsonable(v) for k, v in value.items()}
    if hasattr(value, "value"):
        return value.value
    return value


def _date_from_any(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    try:
        return date.fromisoformat(str(value)[:10])
    except Exception:
        return None


def _field_type_for_storage(
    field_type: Union[str, InventoryFieldType, None],
    *,
    key: Optional[str] = None,
    category: Optional[Union[str, InventoryFieldCategory]] = None,
    unit: Optional[str] = None,
) -> InventoryFieldType:
    """Return a DB-safe InventoryFieldType.

    Older ranch templates used a pseudo field type named SYSTEM_CALCULATED to
    mean "this is an output calculated by the system". PostgreSQL deployments
    that were created before that enum value existed cannot store it. To keep
    the schema stable, calculated output fields are stored as normal concrete
    types (number, currency, percentage, text) while field_direction=OUTPUT and
    is_system_calculated=True carry the calculation meaning.
    """
    raw = _enum_value(field_type).lower() if field_type is not None else ""
    if raw in {"system_calculated", "system", "calculated", "output"}:
        key_l = str(key or "").lower()
        category_l = _enum_value(category).lower() if category is not None else ""
        unit_l = str(unit or "").lower()
        if "rate" in key_l or "percentage" in key_l or unit_l in {"%", "percent", "percentage"}:
            return InventoryFieldType.PERCENTAGE
        if "status" in key_l:
            return InventoryFieldType.TEXT
        if category_l in {"cost", "sales"} or "tzs" in unit_l or "amount" in key_l or "value" in key_l or "cost" in key_l or "price" in key_l or "profit" in key_l:
            return InventoryFieldType.CURRENCY
        return InventoryFieldType.NUMBER
    return _to_enum(InventoryFieldType, field_type, InventoryFieldType.TEXT)


def _http_error(status_code: int, detail: str):
    if HTTPException is Exception:  # pragma: no cover
        raise Exception(detail)
    raise HTTPException(status_code=status_code, detail=detail)


def _field(
    name: str,
    key: str,
    field_type: Union[str, InventoryFieldType],
    *,
    direction: Union[str, InventoryFieldDirection] = InventoryFieldDirection.INPUT,
    category: Union[str, InventoryFieldCategory] = InventoryFieldCategory.GENERAL,
    lookup_group: Optional[Union[str, InventoryLookupGroup]] = None,
    unit: Optional[str] = None,
    required: bool = False,
    dashboard: bool = False,
    searchable: bool = False,
    filterable: bool = False,
    sortable: bool = False,
    calculation_key: Optional[str] = None,
    order: int = 0,
    options: Optional[List[Any]] = None,
) -> Dict[str, Any]:
    direction_enum = _to_enum(InventoryFieldDirection, direction, InventoryFieldDirection.INPUT)
    field_type_enum = _field_type_for_storage(field_type, key=key, category=category, unit=unit)
    return {
        "field_name": name,
        "field_key": key,
        "field_code": None,
        "field_type": field_type_enum,
        "field_direction": direction_enum,
        "field_category": _to_enum(InventoryFieldCategory, category, InventoryFieldCategory.GENERAL),
        "lookup_group": _to_enum(InventoryLookupGroup, lookup_group, None) if lookup_group else None,
        "default_value": None,
        "unit_label": unit,
        "is_required": required,
        "is_unique": False,
        "is_searchable": searchable,
        "is_filterable": filterable,
        "is_sortable": sortable,
        "is_indexed": filterable or sortable or searchable,
        "is_report_visible": True,
        "is_dashboard_visible": dashboard,
        "is_system_calculated": direction_enum == InventoryFieldDirection.OUTPUT or _enum_value(field_type).lower() in {"system_calculated", "system", "calculated", "output"},
        "is_user_editable": direction_enum != InventoryFieldDirection.OUTPUT and _enum_value(field_type).lower() not in {"system_calculated", "system", "calculated", "output"},
        "calculation_key": calculation_key,
        "options_json": options or [],
        "validation_json": {},
        "display_json": {},
        "settings_json": {},
        "order_index": order,
    }


def _rule(key: str, name: str, rule_type: InventoryCalculationRuleType, outputs: Dict[str, str], *, order: int = 0) -> Dict[str, Any]:
    return {
        "rule_key": key,
        "rule_name": name,
        "rule_type": rule_type,
        "scope": InventoryCalculationScope.ROW,
        "inputs_json": {},
        "outputs_json": outputs,
        "expression_label": None,
        "rule_config_json": {},
        "is_active": True,
        "run_on_create": True,
        "run_on_update": True,
        "run_before_report": True,
        "order_index": order,
    }


def _metric(key: str, label: str, metric_type: InventoryMetricType, *, field_key: Optional[str] = None, unit: Optional[str] = None, order: int = 0) -> Dict[str, Any]:
    return {
        "metric_key": key,
        "metric_type": metric_type,
        "label": label,
        "field_key": field_key,
        "calculation_rule_key": None,
        "unit_label": unit,
        "is_enabled": True,
        "is_report_visible": True,
        "is_dashboard_visible": True,
        "settings_json": {},
        "order_index": order,
    }


class RanchTemplateRegistry:
    """Built-in ranch templates. Users enter inputs; system calculates outputs."""

    @staticmethod
    def all_templates() -> List[Dict[str, Any]]:
        return [
            RanchTemplateRegistry.animal_stock("animals_goat_inventory", "Goat Inventory", InventoryTemplateType.GOAT_INVENTORY, ["Goat"]),
            RanchTemplateRegistry.animal_stock("animals_cattle_inventory", "Cattle Inventory", InventoryTemplateType.CATTLE_INVENTORY, ["Cattle"]),
            RanchTemplateRegistry.feed_inventory(),
            RanchTemplateRegistry.crop_stock(),
            RanchTemplateRegistry.harvest_records(),
            RanchTemplateRegistry.fertilizer_inventory(),
            RanchTemplateRegistry.fuel_usage(),
            RanchTemplateRegistry.spare_parts_inventory(),
            RanchTemplateRegistry.service_records(),
        ]

    @staticmethod
    def animal_stock(template_key: str, name: str, inventory_type: InventoryTemplateType, animal_options: List[str]) -> Dict[str, Any]:
        fields = [
            _field("Animal Type", "animal_type", "dropdown", category="classification", lookup_group="animal_type", required=True, filterable=True, order=1, options=animal_options),
            _field("Breed", "breed", "dropdown", category="classification", lookup_group="animal_breed", filterable=True, order=2),
            _field("Animal Category", "animal_category", "dropdown", category="classification", lookup_group="animal_category", filterable=True, order=3, options=["Adult Male", "Adult Female", "Young Male", "Young Female", "Newborn", "Other"]),
            _field("Opening Balance", "opening_balance", "number", category="stock", unit="heads", required=True, dashboard=True, sortable=True, order=4),
            _field("Born", "born", "number", category="stock", unit="heads", dashboard=True, order=5),
            _field("Bought", "bought", "number", category="stock", unit="heads", dashboard=True, order=6),
            _field("Transferred In", "transferred_in", "number", category="stock", unit="heads", order=7),
            _field("Sold", "sold", "number", category="sales", unit="heads", dashboard=True, order=8),
            _field("Died", "died", "number", category="health", unit="heads", dashboard=True, order=9),
            _field("Transferred Out", "transferred_out", "number", category="stock", unit="heads", order=10),
            _field("Culled", "culled", "number", category="stock", unit="heads", order=11),
            _field("Selling Price", "selling_price", "currency", category="sales", unit="TZS", order=12),
            _field("Purchase Price", "purchase_price", "currency", category="cost", unit="TZS", order=13),
            _field("Current Balance", "current_balance", "system_calculated", direction="output", category="stock", unit="heads", calculation_key="animal_current_balance", dashboard=True, order=14),
            _field("Total Additions", "total_additions", "system_calculated", direction="output", category="stock", unit="heads", calculation_key="animal_total_additions", order=15),
            _field("Total Reductions", "total_reductions", "system_calculated", direction="output", category="stock", unit="heads", calculation_key="animal_total_reductions", order=16),
            _field("Net Movement", "net_movement", "system_calculated", direction="output", category="stock", unit="heads", calculation_key="animal_net_movement", order=17),
            _field("Mortality Rate", "mortality_rate", "system_calculated", direction="output", category="health", unit="%", calculation_key="animal_mortality_rate", order=18),
            _field("Sales Value", "sales_value", "system_calculated", direction="output", category="sales", unit="TZS", calculation_key="animal_sales_value", order=19),
            _field("Purchase Value", "purchase_value", "system_calculated", direction="output", category="cost", unit="TZS", calculation_key="animal_purchase_value", order=20),
            _field("Notes", "notes", "long_text", order=21),
        ]
        rules = [
            _rule("animal_current_balance", "Current Balance", _calc_type("ANIMAL_CURRENT_BALANCE"), {"current_balance": "current_balance"}, order=1),
            _rule("animal_total_additions", "Total Additions", _calc_type("ANIMAL_TOTAL_ADDITIONS"), {"total_additions": "total_additions"}, order=2),
            _rule("animal_total_reductions", "Total Reductions", _calc_type("ANIMAL_TOTAL_REDUCTIONS"), {"total_reductions": "total_reductions"}, order=3),
            _rule("animal_net_movement", "Net Movement", _calc_type("ANIMAL_NET_MOVEMENT"), {"net_movement": "net_movement"}, order=4),
            _rule("animal_mortality_rate", "Mortality Rate", _calc_type("ANIMAL_MORTALITY_RATE"), {"mortality_rate": "mortality_rate"}, order=5),
            _rule("animal_sales_value", "Sales Value", _calc_type("ANIMAL_SALES_VALUE"), {"sales_value": "sales_value"}, order=6),
            _rule("animal_purchase_value", "Purchase Value", _calc_type("ANIMAL_PURCHASE_VALUE"), {"purchase_value": "purchase_value"}, order=7),
        ]
        metrics = [
            _metric("total_current_balance", "Current Balance", InventoryMetricType.SUM, field_key="current_balance", unit="heads", order=1),
            _metric("total_born", "Total Born", InventoryMetricType.SUM, field_key="born", unit="heads", order=2),
            _metric("total_bought", "Total Bought", InventoryMetricType.SUM, field_key="bought", unit="heads", order=3),
            _metric("total_sold", "Total Sold", InventoryMetricType.SUM, field_key="sold", unit="heads", order=4),
            _metric("total_died", "Total Died", InventoryMetricType.SUM, field_key="died", unit="heads", order=5),
        ]
        return {"template_key": template_key, "name": name, "description": "Daily animal movement and balance tracking.", "department": RanchDepartment.ANIMALS, "inventory_type": inventory_type, "fields": fields, "rules": rules, "metrics": metrics}

    @staticmethod
    def feed_inventory() -> Dict[str, Any]:
        fields = [
            _field("Feed Type", "feed_type", "dropdown", category="classification", lookup_group="feed_type", required=True, filterable=True, order=1),
            _field("Opening Stock", "opening_stock", "number", category="stock", unit="kg", required=True, dashboard=True, order=2),
            _field("Purchased", "purchased", "number", category="stock", unit="kg", order=3),
            _field("Produced", "produced", "number", category="stock", unit="kg", order=4),
            _field("Used", "used", "number", category="stock", unit="kg", order=5),
            _field("Damaged", "damaged", "number", category="stock", unit="kg", order=6),
            _field("Unit Cost", "unit_cost", "currency", category="cost", unit="TZS", order=7),
            _field("Supplier", "supplier", "dropdown", category="classification", lookup_group="supplier", order=8),
            _field("Closing Stock", "closing_stock", "system_calculated", direction="output", category="stock", unit="kg", calculation_key="feed_closing_stock", dashboard=True, order=9),
            _field("Total Feed Cost", "total_feed_cost", "system_calculated", direction="output", category="cost", unit="TZS", calculation_key="feed_total_cost", order=10),
            _field("Estimated Days Remaining", "estimated_days_remaining", "system_calculated", direction="output", category="stock", unit="days", calculation_key="feed_days_remaining", order=11),
            _field("Notes", "notes", "long_text", order=12),
        ]
        rules = [
            _rule("feed_closing_stock", "Closing Feed Stock", _calc_type("FEED_CLOSING_STOCK"), {"closing_stock": "closing_stock"}, order=1),
            _rule("feed_total_cost", "Total Feed Cost", _calc_type("FEED_TOTAL_COST"), {"total_feed_cost": "total_feed_cost"}, order=2),
            _rule("feed_days_remaining", "Estimated Days Remaining", _calc_type("FEED_DAYS_REMAINING"), {"estimated_days_remaining": "estimated_days_remaining"}, order=3),
        ]
        metrics = [
            _metric("total_closing_stock", "Closing Feed Stock", InventoryMetricType.SUM, field_key="closing_stock", unit="kg", order=1),
            _metric("total_used", "Feed Used", InventoryMetricType.SUM, field_key="used", unit="kg", order=2),
            _metric("total_feed_cost", "Total Feed Cost", InventoryMetricType.SUM, field_key="total_feed_cost", unit="TZS", order=3),
        ]
        return {"template_key": "animals_feed_inventory", "name": "Feed Inventory", "description": "Feed stock balances and cost tracking.", "department": RanchDepartment.ANIMALS, "inventory_type": InventoryTemplateType.FEED_INVENTORY, "fields": fields, "rules": rules, "metrics": metrics}

    @staticmethod
    def crop_stock() -> Dict[str, Any]:
        fields = [
            _field("Crop Type", "crop_type", "dropdown", category="classification", lookup_group="crop_type", required=True, filterable=True, order=1),
            _field("Crop Variety", "crop_variety", "dropdown", category="classification", lookup_group="crop_variety", filterable=True, order=2),
            _field("Field / Plot", "field_or_plot", "dropdown", category="location", lookup_group="field_or_plot", filterable=True, order=3),
            _field("Opening Stock", "opening_stock", "number", category="stock", unit="kg", required=True, dashboard=True, order=4),
            _field("Harvested", "harvested", "number", category="production", unit="kg", order=5),
            _field("Purchased", "purchased", "number", category="stock", unit="kg", order=6),
            _field("Used", "used", "number", category="stock", unit="kg", order=7),
            _field("Sold", "sold", "number", category="sales", unit="kg", order=8),
            _field("Damaged", "damaged", "number", category="stock", unit="kg", order=9),
            _field("Unit Price", "unit_price", "currency", category="sales", unit="TZS", order=10),
            _field("Closing Stock", "closing_stock", "system_calculated", direction="output", category="stock", unit="kg", calculation_key="crop_closing_stock", dashboard=True, order=11),
            _field("Sales Value", "sales_value", "system_calculated", direction="output", category="sales", unit="TZS", calculation_key="crop_estimated_sales_value", order=12),
            _field("Notes", "notes", "long_text", order=13),
        ]
        rules = [
            _rule("crop_closing_stock", "Closing Crop Stock", _calc_type("CROP_CLOSING_STOCK"), {"closing_stock": "closing_stock"}, order=1),
            _rule("crop_sales_value", "Crop Sales Value", _calc_type("CROP_ESTIMATED_SALES_VALUE"), {"sales_value": "sales_value"}, order=2),
        ]
        metrics = [
            _metric("total_closing_stock", "Closing Stock", InventoryMetricType.SUM, field_key="closing_stock", unit="kg", order=1),
            _metric("total_harvested", "Total Harvested", InventoryMetricType.SUM, field_key="harvested", unit="kg", order=2),
            _metric("total_sold", "Total Sold", InventoryMetricType.SUM, field_key="sold", unit="kg", order=3),
        ]
        return {"template_key": "crops_crop_stock", "name": "Crop Stock", "description": "Crop stock inputs and automatic closing stock.", "department": RanchDepartment.CROPS, "inventory_type": InventoryTemplateType.CROP_STOCK, "fields": fields, "rules": rules, "metrics": metrics}

    @staticmethod
    def harvest_records() -> Dict[str, Any]:
        fields = [
            _field("Crop Type", "crop_type", "dropdown", category="classification", lookup_group="crop_type", required=True, filterable=True, order=1),
            _field("Crop Variety", "crop_variety", "dropdown", category="classification", lookup_group="crop_variety", order=2),
            _field("Field / Plot", "field_or_plot", "dropdown", category="location", lookup_group="field_or_plot", order=3),
            _field("Acres Planted", "acres_planted", "number", category="production", unit="acres", order=4),
            _field("Harvest Quantity", "harvest_quantity", "number", category="production", unit="kg", required=True, order=5),
            _field("Labour Cost", "labour_cost", "currency", category="cost", unit="TZS", order=6),
            _field("Transport Cost", "transport_cost", "currency", category="cost", unit="TZS", order=7),
            _field("Packaging Cost", "packaging_cost", "currency", category="cost", unit="TZS", order=8),
            _field("Other Cost", "other_cost", "currency", category="cost", unit="TZS", order=9),
            _field("Selling Price", "selling_price", "currency", category="sales", unit="TZS/kg", order=10),
            _field("Yield Per Acre", "yield_per_acre", "system_calculated", direction="output", category="production", unit="kg/acre", calculation_key="crop_yield_per_acre", order=11),
            _field("Total Production Cost", "total_production_cost", "system_calculated", direction="output", category="cost", unit="TZS", calculation_key="crop_total_production_cost", order=12),
            _field("Estimated Sales Value", "estimated_sales_value", "system_calculated", direction="output", category="sales", unit="TZS", calculation_key="crop_estimated_sales_value", order=13),
            _field("Profit / Loss", "profit_loss", "system_calculated", direction="output", category="sales", unit="TZS", calculation_key="crop_profit_loss", order=14),
            _field("Cost Per Unit", "cost_per_unit", "system_calculated", direction="output", category="cost", unit="TZS/kg", calculation_key="crop_cost_per_unit", order=15),
            _field("Notes", "notes", "long_text", order=16),
        ]
        rules = [
            _rule("crop_yield_per_acre", "Yield Per Acre", _calc_type("CROP_YIELD_PER_ACRE"), {"yield_per_acre": "yield_per_acre"}, order=1),
            _rule("crop_total_production_cost", "Total Production Cost", _calc_type("CROP_TOTAL_PRODUCTION_COST"), {"total_production_cost": "total_production_cost"}, order=2),
            _rule("crop_estimated_sales_value", "Estimated Sales Value", _calc_type("CROP_ESTIMATED_SALES_VALUE"), {"estimated_sales_value": "estimated_sales_value"}, order=3),
            _rule("crop_profit_loss", "Profit / Loss", _calc_type("CROP_PROFIT_LOSS"), {"profit_loss": "profit_loss"}, order=4),
            _rule("crop_cost_per_unit", "Cost Per Unit", _calc_type("CROP_COST_PER_UNIT"), {"cost_per_unit": "cost_per_unit"}, order=5),
        ]
        metrics = [
            _metric("total_harvest_quantity", "Harvest Quantity", InventoryMetricType.SUM, field_key="harvest_quantity", unit="kg", order=1),
            _metric("avg_yield_per_acre", "Average Yield Per Acre", InventoryMetricType.AVERAGE, field_key="yield_per_acre", unit="kg/acre", order=2),
            _metric("total_profit_loss", "Profit / Loss", InventoryMetricType.SUM, field_key="profit_loss", unit="TZS", order=3),
        ]
        return {"template_key": "crops_harvest_records", "name": "Harvest Records", "description": "Harvest, yield, cost, and profit tracking.", "department": RanchDepartment.CROPS, "inventory_type": InventoryTemplateType.HARVEST_RECORDS, "fields": fields, "rules": rules, "metrics": metrics}

    @staticmethod
    def fertilizer_inventory() -> Dict[str, Any]:
        base = RanchTemplateRegistry.feed_inventory()
        base.update({"template_key": "crops_fertilizer_inventory", "name": "Fertilizer Inventory", "department": RanchDepartment.CROPS, "inventory_type": InventoryTemplateType.FERTILIZER_INVENTORY})
        for f in base["fields"]:
            if f["field_key"] == "feed_type":
                f.update({"field_name": "Fertilizer Type", "field_key": "fertilizer_type", "lookup_group": InventoryLookupGroup.FERTILIZER_TYPE})
            if f.get("calculation_key") == "feed_closing_stock":
                f["calculation_key"] = "fertilizer_closing_stock"
        base["rules"] = [_rule("fertilizer_closing_stock", "Closing Fertilizer Stock", _calc_type("FERTILIZER_CLOSING_STOCK"), {"closing_stock": "closing_stock"}, order=1)]
        return base

    @staticmethod
    def fuel_usage() -> Dict[str, Any]:
        fields = [
            _field("Machine Type", "machine_type", "dropdown", category="classification", lookup_group="machine_type", required=True, filterable=True, order=1),
            _field("Machine Name", "machine_name", "dropdown", category="classification", lookup_group="machine_name", required=True, searchable=True, filterable=True, order=2),
            _field("Registration / Asset Code", "asset_code", "text", category="classification", searchable=True, order=3),
            _field("Operator", "operator", "dropdown", category="classification", lookup_group="operator", order=4),
            _field("Opening Hours", "opening_hours", "number", category="maintenance", unit="hours", order=5),
            _field("Closing Hours", "closing_hours", "number", category="maintenance", unit="hours", order=6),
            _field("Fuel Litres", "fuel_litres", "number", category="quantity", unit="litres", required=True, dashboard=True, order=7),
            _field("Price Per Litre", "price_per_litre", "currency", category="cost", unit="TZS", required=True, order=8),
            _field("Fuel Station", "fuel_station", "text", category="general", order=9),
            _field("Receipt Number", "receipt_number", "text", category="general", searchable=True, order=10),
            _field("Running Hours", "running_hours", "system_calculated", direction="output", category="maintenance", unit="hours", calculation_key="machine_running_hours", dashboard=True, order=11),
            _field("Total Fuel Cost", "total_fuel_cost", "system_calculated", direction="output", category="cost", unit="TZS", calculation_key="fuel_total_cost", dashboard=True, order=12),
            _field("Fuel Per Hour", "fuel_per_hour", "system_calculated", direction="output", category="quantity", unit="litres/hour", calculation_key="fuel_per_hour", order=13),
            _field("Notes", "notes", "long_text", order=14),
        ]
        rules = [
            _rule("machine_running_hours", "Running Hours", _calc_type("MACHINE_RUNNING_HOURS"), {"running_hours": "running_hours"}, order=1),
            _rule("fuel_total_cost", "Total Fuel Cost", _calc_type("FUEL_TOTAL_COST"), {"total_fuel_cost": "total_fuel_cost"}, order=2),
            _rule("fuel_per_hour", "Fuel Per Hour", _calc_type("FUEL_PER_HOUR"), {"fuel_per_hour": "fuel_per_hour"}, order=3),
        ]
        metrics = [
            _metric("total_fuel_litres", "Fuel Litres", InventoryMetricType.SUM, field_key="fuel_litres", unit="litres", order=1),
            _metric("total_fuel_cost", "Total Fuel Cost", InventoryMetricType.SUM, field_key="total_fuel_cost", unit="TZS", order=2),
        ]
        return {"template_key": "machinery_fuel_usage", "name": "Fuel Usage", "description": "Fuel, cost, and running hours tracking.", "department": RanchDepartment.MACHINERY, "inventory_type": InventoryTemplateType.FUEL_USAGE, "fields": fields, "rules": rules, "metrics": metrics}

    @staticmethod
    def spare_parts_inventory() -> Dict[str, Any]:
        fields = [
            _field("Part Name", "part_name", "text", category="classification", required=True, searchable=True, order=1),
            _field("Part Code", "part_code", "text", category="classification", searchable=True, order=2),
            _field("Machine Type", "machine_type", "dropdown", category="classification", lookup_group="machine_type", filterable=True, order=3),
            _field("Machine Name", "machine_name", "dropdown", category="classification", lookup_group="machine_name", order=4),
            _field("Opening Stock", "opening_stock", "number", category="stock", required=True, order=5),
            _field("Purchased", "purchased", "number", category="stock", order=6),
            _field("Used", "used", "number", category="stock", order=7),
            _field("Damaged", "damaged", "number", category="stock", order=8),
            _field("Unit Cost", "unit_cost", "currency", category="cost", unit="TZS", order=9),
            _field("Supplier", "supplier", "dropdown", category="classification", lookup_group="supplier", order=10),
            _field("Closing Stock", "closing_stock", "system_calculated", direction="output", category="stock", calculation_key="spare_parts_closing_stock", dashboard=True, order=11),
            _field("Remaining Stock Value", "remaining_stock_value", "system_calculated", direction="output", category="cost", unit="TZS", calculation_key="spare_parts_remaining_value", order=12),
            _field("Notes", "notes", "long_text", order=13),
        ]
        rules = [
            _rule("spare_parts_closing_stock", "Closing Stock", _calc_type("SPARE_PARTS_CLOSING_STOCK"), {"closing_stock": "closing_stock"}, order=1),
            _rule("spare_parts_remaining_value", "Remaining Stock Value", _calc_type("SPARE_PARTS_REMAINING_VALUE"), {"remaining_stock_value": "remaining_stock_value"}, order=2),
        ]
        metrics = [
            _metric("total_closing_stock", "Closing Stock", InventoryMetricType.SUM, field_key="closing_stock", order=1),
            _metric("remaining_stock_value", "Remaining Stock Value", InventoryMetricType.SUM, field_key="remaining_stock_value", unit="TZS", order=2),
        ]
        return {"template_key": "machinery_spare_parts_inventory", "name": "Spare Parts Inventory", "description": "Spare parts stock and value tracking.", "department": RanchDepartment.MACHINERY, "inventory_type": InventoryTemplateType.SPARE_PARTS_INVENTORY, "fields": fields, "rules": rules, "metrics": metrics}

    @staticmethod
    def service_records() -> Dict[str, Any]:
        fields = [
            _field("Machine Type", "machine_type", "dropdown", category="classification", lookup_group="machine_type", required=True, order=1),
            _field("Machine Name", "machine_name", "dropdown", category="classification", lookup_group="machine_name", required=True, order=2),
            _field("Service Type", "service_type", "text", category="maintenance", order=3),
            _field("Problem Description", "problem_description", "long_text", category="maintenance", order=4),
            _field("Labour Cost", "labour_cost", "currency", category="cost", unit="TZS", order=5),
            _field("Parts Cost", "parts_cost", "currency", category="cost", unit="TZS", order=6),
            _field("Other Cost", "other_cost", "currency", category="cost", unit="TZS", order=7),
            _field("Technician", "technician", "text", category="maintenance", order=8),
            _field("Next Service Date", "next_service_date", "date", category="maintenance", order=9),
            _field("Total Maintenance Cost", "total_maintenance_cost", "system_calculated", direction="output", category="cost", unit="TZS", calculation_key="maintenance_total_cost", order=10),
            _field("Service Due Status", "service_due_status", "system_calculated", direction="output", category="maintenance", calculation_key="service_due_status", order=11),
            _field("Notes", "notes", "long_text", order=12),
        ]
        rules = [
            _rule("maintenance_total_cost", "Total Maintenance Cost", _calc_type("MAINTENANCE_TOTAL_COST"), {"total_maintenance_cost": "total_maintenance_cost"}, order=1),
            _rule("service_due_status", "Service Due Status", _calc_type("SERVICE_DUE_STATUS"), {"service_due_status": "service_due_status"}, order=2),
        ]
        metrics = [_metric("total_maintenance_cost", "Maintenance Cost", InventoryMetricType.SUM, field_key="total_maintenance_cost", unit="TZS", order=1)]
        return {"template_key": "machinery_service_records", "name": "Service Records", "description": "Machinery service and maintenance cost tracking.", "department": RanchDepartment.MACHINERY, "inventory_type": InventoryTemplateType.SERVICE_RECORDS, "fields": fields, "rules": rules, "metrics": metrics}


class DynamicInventoryService:
    """Ranch inventory service: templates, daily records, calculations, reports."""

    pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto") if CryptContext else None

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @classmethod
    def _hash_password(cls, password: str) -> str:
        if cls.pwd_context:
            return cls.pwd_context.hash(password)
        return hashlib.sha256(password.encode("utf-8")).hexdigest()

    @classmethod
    def _coerce_audit_action(cls, action: Any) -> InventoryAuditAction:
        """Return a PostgreSQL/SQLAlchemy-safe audit enum member.

        The model uses ``SAEnum(InventoryAuditAction)``. PostgreSQL stores the
        enum MEMBER NAMES by default, for example ``CREATE_INVENTORY``. Passing
        a raw lowercase string such as ``create_inventory`` makes psycopg2 send
        that lowercase label directly and PostgreSQL rejects it.

        This helper accepts either enum members, enum names, or public values and
        always returns the matching ``InventoryAuditAction`` member so SQLAlchemy
        can bind the correct database enum label.
        """
        if isinstance(action, InventoryAuditAction):
            return action

        raw = str(action or "system_calculation").strip()
        if "." in raw:
            raw = raw.split(".")[-1]

        normalized = raw.replace("-", "_").replace(" ", "_")
        for item in InventoryAuditAction:
            if normalized.upper() == item.name:
                return item
            if normalized.lower() == str(item.value).lower():
                return item

        return InventoryAuditAction.SYSTEM_CALCULATION

    @classmethod
    def _audit(cls, db: Session, *, action: Any, actor_user_id: Optional[int] = None, inventory_id: Optional[int] = None, period_id: Optional[int] = None, row_id: Optional[int] = None, field_id: Optional[int] = None, description: Optional[str] = None, old: Optional[Dict[str, Any]] = None, new: Optional[Dict[str, Any]] = None, context: Optional[Dict[str, Any]] = None) -> DynamicInventoryAuditLog:
        """Create an audit log without breaking PostgreSQL enum binding."""
        log = DynamicInventoryAuditLog(
            inventory_id=inventory_id,
            period_id=period_id,
            row_id=row_id,
            field_id=field_id,
            actor_user_id=actor_user_id,
            action=cls._coerce_audit_action(action),
            description=description,
            old_value_json=_jsonable(old) if old is not None else None,
            new_value_json=_jsonable(new) if new is not None else None,
            context_json=_jsonable(context or {}),
        )
        db.add(log)
        return log

    @classmethod
    def _unique_slug(cls, db: Session, title: str, inventory_id: Optional[int] = None) -> str:
        base = _slugify(title)
        slug = base
        suffix = 2
        while True:
            query = db.query(DynamicInventory).filter(DynamicInventory.slug == slug)
            if inventory_id:
                query = query.filter(DynamicInventory.id != inventory_id)
            if not query.first():
                return slug
            slug = f"{base}-{suffix}"
            suffix += 1

    @staticmethod
    def _template_field_from_spec(template_id: int, spec: Dict[str, Any], index: int, actor_user_id: Optional[int]) -> DynamicInventoryTemplateField:
        return DynamicInventoryTemplateField(
            template_id=template_id,
            field_name=spec["field_name"],
            field_key=spec["field_key"],
            field_code=spec.get("field_code") or _field_code(index),
            field_type=_field_type_for_storage(spec.get("field_type"), key=spec.get("field_key"), category=spec.get("field_category"), unit=spec.get("unit_label")),
            field_direction=_to_enum(InventoryFieldDirection, spec.get("field_direction"), InventoryFieldDirection.INPUT),
            field_category=_to_enum(InventoryFieldCategory, spec.get("field_category"), InventoryFieldCategory.GENERAL),
            lookup_group=_to_enum(InventoryLookupGroup, spec.get("lookup_group"), None),
            description=spec.get("description"),
            placeholder=spec.get("placeholder"),
            default_value=spec.get("default_value"),
            unit_label=spec.get("unit_label"),
            is_required=bool(spec.get("is_required", False)),
            is_unique=bool(spec.get("is_unique", False)),
            is_searchable=bool(spec.get("is_searchable", False)),
            is_filterable=bool(spec.get("is_filterable", False)),
            is_sortable=bool(spec.get("is_sortable", False)),
            is_indexed=bool(spec.get("is_indexed", False)),
            is_report_visible=bool(spec.get("is_report_visible", True)),
            is_dashboard_visible=bool(spec.get("is_dashboard_visible", False)),
            is_system_calculated=bool(spec.get("is_system_calculated", False)),
            is_user_editable=bool(spec.get("is_user_editable", not spec.get("is_system_calculated", False))),
            calculation_key=spec.get("calculation_key"),
            options_json=_jsonable(spec.get("options_json") or []),
            validation_json=_jsonable(spec.get("validation_json") or {}),
            display_json=_jsonable(spec.get("display_json") or {}),
            settings_json=_jsonable(spec.get("settings_json") or {}),
            order_index=spec.get("order_index", index + 1),
            created_by_user_id=actor_user_id,
            updated_by_user_id=actor_user_id,
        )

    @staticmethod
    def _template_rule_from_spec(template_id: int, spec: Dict[str, Any], index: int, actor_user_id: Optional[int]) -> DynamicInventoryTemplateCalculationRule:
        return DynamicInventoryTemplateCalculationRule(
            template_id=template_id,
            rule_key=spec["rule_key"],
            rule_name=spec["rule_name"],
            rule_type=_to_enum(InventoryCalculationRuleType, spec.get("rule_type"), _calc_type("CUSTOM_SYSTEM_RULE")),
            scope=_to_enum(InventoryCalculationScope, spec.get("scope"), InventoryCalculationScope.ROW),
            inputs_json=_jsonable(spec.get("inputs_json") or {}),
            outputs_json=_jsonable(spec.get("outputs_json") or {}),
            expression_label=spec.get("expression_label"),
            rule_config_json=_jsonable(spec.get("rule_config_json") or {}),
            is_active=bool(spec.get("is_active", True)),
            run_on_create=bool(spec.get("run_on_create", True)),
            run_on_update=bool(spec.get("run_on_update", True)),
            run_before_report=bool(spec.get("run_before_report", True)),
            order_index=spec.get("order_index", index + 1),
            created_by_user_id=actor_user_id,
            updated_by_user_id=actor_user_id,
        )

    @staticmethod
    def _template_metric_from_spec(template_id: int, spec: Dict[str, Any], index: int, actor_user_id: Optional[int]) -> DynamicInventoryTemplateMetric:
        return DynamicInventoryTemplateMetric(
            template_id=template_id,
            metric_key=spec["metric_key"],
            metric_type=_to_enum(InventoryMetricType, spec.get("metric_type"), InventoryMetricType.SUM),
            label=spec["label"],
            field_key=spec.get("field_key"),
            calculation_rule_key=spec.get("calculation_rule_key"),
            unit_label=spec.get("unit_label"),
            is_enabled=bool(spec.get("is_enabled", True)),
            is_report_visible=bool(spec.get("is_report_visible", True)),
            is_dashboard_visible=bool(spec.get("is_dashboard_visible", True)),
            settings_json=_jsonable(spec.get("settings_json") or {}),
            order_index=spec.get("order_index", index + 1),
            created_by_user_id=actor_user_id,
            updated_by_user_id=actor_user_id,
        )

    @staticmethod
    def _cell_value(cell: DynamicInventoryValue) -> Any:
        if cell.value_number is not None:
            return cell.value_number
        if cell.value_date is not None:
            return cell.value_date
        if cell.value_boolean is not None:
            return cell.value_boolean
        if cell.value_json is not None:
            return cell.value_json
        return cell.value_text

    @classmethod
    def _row_values(cls, row: DynamicInventoryRow, fields: Sequence[DynamicInventoryField]) -> Dict[str, Any]:
        by_id = {f.id: f for f in fields}
        values: Dict[str, Any] = {}
        for cell in row.values:
            field = by_id.get(cell.field_id)
            if field:
                values[field.field_key] = _jsonable(cls._cell_value(cell))
        return values

    @classmethod
    def _row_out(cls, row: DynamicInventoryRow, fields: Sequence[DynamicInventoryField]) -> Dict[str, Any]:
        """Serialize a row safely across older/newer model versions.

        Some existing databases/models do not have optional entity columns such as
        primary_entity_code. The user-facing entity details are also available in
        values/metadata, so never fail response serialization because one optional
        attribute is absent.
        """
        values = cls._row_values(row, fields)
        metadata = getattr(row, "metadata_json", None) or {}
        return {
            "id": row.id,
            "inventory_id": row.inventory_id,
            "period_id": row.period_id,
            "row_number": row.row_number,
            "row_label": row.row_label,
            "record_date": row.record_date,
            "primary_entity_type": getattr(row, "primary_entity_type", None) or metadata.get("primary_entity_type") or values.get("animal_type") or values.get("crop_type") or values.get("machine_type") or values.get("feed_type") or values.get("part_name"),
            "primary_entity_name": getattr(row, "primary_entity_name", None) or metadata.get("primary_entity_name") or values.get("breed") or values.get("crop_variety") or values.get("machine_name") or values.get("part_name") or values.get("field_or_plot"),
            "primary_entity_code": getattr(row, "primary_entity_code", None) or metadata.get("primary_entity_code") or values.get("asset_code") or values.get("part_code") or values.get("receipt_number") or values.get("tag_number"),
            "is_deleted": row.is_deleted,
            "computed_json": row.computed_json or {},
            "metadata_json": metadata,
            "values": values,
            "created_at": row.created_at,
            "updated_at": row.updated_at,
        }

    @classmethod
    def _period_out(cls, period: DynamicInventoryPeriod) -> Dict[str, Any]:
        return {
            "id": period.id,
            "inventory_id": period.inventory_id,
            "period_type": period.period_type,
            "period_date": period.period_date,
            "start_date": period.start_date,
            "end_date": period.end_date,
            "title": period.title,
            "status": period.status,
            "notes": period.notes,
            "rejection_reason": period.rejection_reason,
            "submitted_at": period.submitted_at,
            "approved_at": period.approved_at,
            "rejected_at": period.rejected_at,
            "locked_at": period.locked_at,
            "summary_json": period.summary_json or {},
            "settings_json": period.settings_json or {},
            "created_at": period.created_at,
            "updated_at": period.updated_at,
        }

    @classmethod
    def _inventory_summary(cls, inventory: DynamicInventory) -> Dict[str, Any]:
        today = date.today()
        periods = list(getattr(inventory, "periods", []) or [])
        today_period = next((p for p in periods if p.period_type == InventoryPeriodType.DAILY and p.period_date == today), None)
        return {
            "id": inventory.id,
            "title": inventory.title,
            "slug": inventory.slug,
            "description": inventory.description,
            "department": inventory.department,
            "inventory_type": inventory.inventory_type,
            "template_id": inventory.template_id,
            "report_title": inventory.report_title,
            "reporter_name": inventory.reporter_name,
            "company_name": inventory.company_name,
            "access_type": inventory.access_type,
            "status": inventory.status,
            "is_active": inventory.is_active,
            "today_period_id": today_period.id if today_period else None,
            "last_period_status": today_period.status if today_period else None,
            "today_row_count": len([r for r in today_period.rows if not r.is_deleted]) if today_period else 0,
            "pending_approvals": sum(1 for p in periods if p.status == InventoryPeriodStatus.SUBMITTED),
            "open_alerts": len([a for a in getattr(inventory, "alerts", []) if a.status == InventoryAlertStatus.OPEN]),
            "created_at": inventory.created_at,
            "updated_at": inventory.updated_at,
        }

    # ------------------------------------------------------------------
    # Seeds and templates
    # ------------------------------------------------------------------

    @classmethod
    def seed_default_lookup_options(cls, db: Session, actor_user_id: Optional[int] = None) -> Dict[str, int]:
        seeds = [
            (RanchDepartment.ANIMALS, InventoryLookupGroup.ANIMAL_TYPE, ["Goat", "Cattle", "Sheep", "Poultry", "Pig", "Fish", "Other"]),
            (RanchDepartment.ANIMALS, InventoryLookupGroup.ANIMAL_CATEGORY, ["Adult Male", "Adult Female", "Young Male", "Young Female", "Newborn", "Other"]),
            (RanchDepartment.ANIMALS, InventoryLookupGroup.ANIMAL_BREED, ["Local", "Boer", "Saanen", "Friesian", "Boran", "Other"]),
            (RanchDepartment.CROPS, InventoryLookupGroup.CROP_TYPE, ["Maize", "Beans", "Rice", "Sunflower", "Vegetables", "Fruits", "Other"]),
            (RanchDepartment.CROPS, InventoryLookupGroup.CROP_VARIETY, ["Local", "Hybrid", "Improved", "Organic", "Other"]),
            (RanchDepartment.CROPS, InventoryLookupGroup.FIELD_OR_PLOT, ["Field A", "Field B", "Field C", "Greenhouse", "Nursery", "Other"]),
            (RanchDepartment.MACHINERY, InventoryLookupGroup.MACHINE_TYPE, ["Tractor", "Truck", "Generator", "Water Pump", "Harvester", "Motorbike", "Other"]),
            (RanchDepartment.MACHINERY, InventoryLookupGroup.MACHINE_NAME, ["Tractor 01", "Truck 01", "Generator 01", "Water Pump 01"]),
            (None, InventoryLookupGroup.FEED_TYPE, ["Hay", "Silage", "Concentrate", "Mineral", "Dairy Meal", "Other"]),
            (None, InventoryLookupGroup.FERTILIZER_TYPE, ["NPK", "Urea", "DAP", "CAN", "Manure", "Compost", "Other"]),
            (None, InventoryLookupGroup.SUPPLIER, ["Internal", "Local Supplier", "External Supplier", "Other"]),
            (None, InventoryLookupGroup.OPERATOR, ["Operator 1", "Operator 2", "Driver 1", "Other"]),
        ]
        created = 0
        skipped = 0
        for department, group, labels in seeds:
            for index, label in enumerate(labels):
                value = _normalize_key(label)
                exists = db.query(DynamicInventoryLookupOption).filter(DynamicInventoryLookupOption.group == group, DynamicInventoryLookupOption.value == value).first()
                if exists:
                    skipped += 1
                    continue
                db.add(DynamicInventoryLookupOption(
                    department=department,
                    group=group,
                    label=label,
                    value=value,
                    is_system_option=True,
                    is_active=True,
                    order_index=index + 1,
                    created_by_user_id=actor_user_id,
                    updated_by_user_id=actor_user_id,
                ))
                created += 1
        db.commit()
        return {"created": created, "skipped": skipped}

    @classmethod
    def seed_default_templates(cls, db: Session, actor_user_id: Optional[int] = None) -> Dict[str, int]:
        created = 0
        updated = 0
        for spec in RanchTemplateRegistry.all_templates():
            template = db.query(DynamicInventoryTemplate).filter(DynamicInventoryTemplate.template_key == spec["template_key"]).first()
            if template:
                template.name = spec["name"]
                template.description = spec.get("description")
                template.department = spec["department"]
                template.inventory_type = spec["inventory_type"]
                template.fields_json = _jsonable(spec.get("fields") or [])
                template.calculation_rules_json = _jsonable(spec.get("rules") or [])
                template.metrics_json = _jsonable(spec.get("metrics") or [])
                template.updated_by_user_id = actor_user_id
                template.template_fields.clear()
                template.template_calculation_rules.clear()
                template.template_metrics.clear()
                updated += 1
            else:
                template = DynamicInventoryTemplate(
                    template_key=spec["template_key"],
                    name=spec["name"],
                    description=spec.get("description"),
                    department=spec["department"],
                    inventory_type=spec["inventory_type"],
                    is_active=True,
                    is_system_template=True,
                    version=1,
                    fields_json=_jsonable(spec.get("fields") or []),
                    calculation_rules_json=_jsonable(spec.get("rules") or []),
                    metrics_json=_jsonable(spec.get("metrics") or []),
                    report_config_json={},
                    dashboard_config_json={},
                    settings_json={},
                    created_by_user_id=actor_user_id,
                    updated_by_user_id=actor_user_id,
                )
                db.add(template)
                created += 1
            db.flush()
            for index, field_spec in enumerate(spec.get("fields") or []):
                template.template_fields.append(cls._template_field_from_spec(template.id, field_spec, index, actor_user_id))
            for index, rule_spec in enumerate(spec.get("rules") or []):
                template.template_calculation_rules.append(DynamicInventoryTemplateCalculationRule(
                    template_id=template.id,
                    rule_key=rule_spec["rule_key"],
                    rule_name=rule_spec["rule_name"],
                    rule_type=rule_spec["rule_type"],
                    scope=rule_spec["scope"],
                    inputs_json=rule_spec.get("inputs_json") or {},
                    outputs_json=rule_spec.get("outputs_json") or {},
                    expression_label=rule_spec.get("expression_label"),
                    rule_config_json=rule_spec.get("rule_config_json") or {},
                    is_active=True,
                    run_on_create=True,
                    run_on_update=True,
                    run_before_report=True,
                    order_index=rule_spec.get("order_index", index + 1),
                    created_by_user_id=actor_user_id,
                    updated_by_user_id=actor_user_id,
                ))
            for index, metric_spec in enumerate(spec.get("metrics") or []):
                template.template_metrics.append(DynamicInventoryTemplateMetric(
                    template_id=template.id,
                    metric_key=metric_spec["metric_key"],
                    metric_type=metric_spec["metric_type"],
                    label=metric_spec["label"],
                    field_key=metric_spec.get("field_key"),
                    calculation_rule_key=metric_spec.get("calculation_rule_key"),
                    unit_label=metric_spec.get("unit_label"),
                    is_enabled=True,
                    is_report_visible=True,
                    is_dashboard_visible=True,
                    settings_json=metric_spec.get("settings_json") or {},
                    order_index=metric_spec.get("order_index", index + 1),
                    created_by_user_id=actor_user_id,
                    updated_by_user_id=actor_user_id,
                ))
        db.commit()
        return {"created": created, "updated": updated}

    @classmethod
    def list_templates(cls, db: Session, department: Optional[Union[str, RanchDepartment]] = None, active_only: bool = True) -> List[DynamicInventoryTemplate]:
        q = db.query(DynamicInventoryTemplate).options(joinedload(DynamicInventoryTemplate.template_fields), joinedload(DynamicInventoryTemplate.template_calculation_rules), joinedload(DynamicInventoryTemplate.template_metrics))
        if department:
            q = q.filter(DynamicInventoryTemplate.department == _to_enum(RanchDepartment, department))
        if active_only:
            q = q.filter(DynamicInventoryTemplate.is_active.is_(True))
        return q.order_by(DynamicInventoryTemplate.department, DynamicInventoryTemplate.name).all()

    @classmethod
    def get_template(cls, db: Session, template_id: int) -> DynamicInventoryTemplate:
        obj = db.query(DynamicInventoryTemplate).options(joinedload(DynamicInventoryTemplate.template_fields), joinedload(DynamicInventoryTemplate.template_calculation_rules), joinedload(DynamicInventoryTemplate.template_metrics)).filter(DynamicInventoryTemplate.id == template_id).first()
        if not obj:
            _http_error(404, "Inventory template not found")
        return obj

    # ------------------------------------------------------------------
    # Inventory creation and CRUD
    # ------------------------------------------------------------------

    @staticmethod
    def _classification_field_from_key(field_key: Any, order: int) -> Dict[str, Any]:
        """Build a real field payload from a classification field key.

        RANCH_CLASSIFICATION_FIELDS stores simple strings such as
        ``animal_type`` and ``field_or_plot``. Those strings cannot be passed to
        ``dict(...)``. Convert them into normal inventory field definitions so
        fallback inventory creation works for every valid inventory type.
        """

        key = _normalize_key(str(field_key or "classification"))
        spec = {
            "animal_type": ("Animal Type", InventoryFieldType.DROPDOWN, InventoryLookupGroup.ANIMAL_TYPE, True),
            "breed": ("Breed", InventoryFieldType.DROPDOWN, InventoryLookupGroup.ANIMAL_BREED, False),
            "animal_category": ("Animal Category", InventoryFieldType.DROPDOWN, InventoryLookupGroup.ANIMAL_CATEGORY, False),
            "crop_type": ("Crop Type", InventoryFieldType.DROPDOWN, InventoryLookupGroup.CROP_TYPE, True),
            "crop_variety": ("Crop Variety", InventoryFieldType.DROPDOWN, InventoryLookupGroup.CROP_VARIETY, False),
            "field_or_plot": ("Field / Plot", InventoryFieldType.DROPDOWN, InventoryLookupGroup.FIELD_OR_PLOT, False),
            "machine_type": ("Machine Type", InventoryFieldType.DROPDOWN, InventoryLookupGroup.MACHINE_TYPE, True),
            "machine_name": ("Machine Name", InventoryFieldType.DROPDOWN, InventoryLookupGroup.MACHINE_NAME, False),
            "asset_code": ("Asset Code", InventoryFieldType.TEXT, None, False),
        }
        label, field_type, lookup_group, required = spec.get(
            key,
            (key.replace("_", " ").title(), InventoryFieldType.TEXT, None, False),
        )
        return _field(
            label,
            key,
            field_type,
            lookup_group=lookup_group,
            required=required,
            searchable=True,
            filterable=True,
            order=order,
        )

    @classmethod
    def _classification_fields_for_department(cls, department: Any, *, start_order: int = 1) -> List[Dict[str, Any]]:
        """Return safe classification field payloads for a department.

        The model may define RANCH_CLASSIFICATION_FIELDS with enum keys or string
        keys depending on the migration/file version, so check both. Each item
        may already be a dict in older/newer versions, or it may be a string key.
        """

        dept_enum = _to_enum(RanchDepartment, department, None)
        dept_value = _enum_value(dept_enum or department)
        raw_items = (
            RANCH_CLASSIFICATION_FIELDS.get(dept_enum)
            or RANCH_CLASSIFICATION_FIELDS.get(dept_value)
            or RANCH_CLASSIFICATION_FIELDS.get(str(dept_value).lower())
            or []
        )

        fields: List[Dict[str, Any]] = []
        for offset, item in enumerate(raw_items):
            order = start_order + offset
            if isinstance(item, dict):
                payload = dict(item)
                payload["order_index"] = payload.get("order_index") or order
                fields.append(payload)
            else:
                fields.append(cls._classification_field_from_key(item, order))
        return fields

    @classmethod
    def _generic_fields_for_inventory(cls, department: RanchDepartment, inventory_type: InventoryTemplateType) -> List[Dict[str, Any]]:
        """Fallback fields used when a specific system template is not seeded yet.

        This prevents the API from returning 404/500 for valid inventory_type enum
        values that are selectable on the frontend but do not yet have a full
        predefined template. The inventory is still created and can later be
        customized or upgraded to a richer template.
        """

        type_label = _enum_value(inventory_type).replace("_", " ").title()
        base_fields = cls._classification_fields_for_department(department, start_order=1)
        order = len(base_fields) + 1

        generic_fields = [
            _field("Record Date", "record_date", InventoryFieldType.DATE, required=True, dashboard=True, filterable=True, sortable=True, order=order),
            _field(f"{type_label} Description", "record_description", InventoryFieldType.TEXT, required=True, searchable=True, order=order + 1),
            _field("Quantity", "quantity", InventoryFieldType.NUMBER, category=InventoryFieldCategory.QUANTITY, dashboard=True, sortable=True, order=order + 2),
            _field("Unit", "unit", InventoryFieldType.TEXT, order=order + 3),
            _field("Amount / Cost", "amount", InventoryFieldType.CURRENCY, category=InventoryFieldCategory.COST, unit="TZS", dashboard=True, sortable=True, order=order + 4),
            _field("Status", "record_status", InventoryFieldType.DROPDOWN, options=["Active", "Pending", "Completed", "Draft"], filterable=True, order=order + 5),
            _field("Notes", "notes", InventoryFieldType.LONG_TEXT, order=order + 6),
        ]
        return base_fields + generic_fields

    @classmethod
    def create_inventory_from_template(cls, db: Session, payload: Any, actor_user_id: Optional[int] = None) -> DynamicInventory:
        data = _dump(payload)

        department = _to_enum(RanchDepartment, data.get("department"), None)
        inv_type = _to_enum(InventoryTemplateType, data.get("inventory_type"), None)
        template = None

        # First try an explicit template id when the frontend has one.
        template_id = data.get("template_id")
        if template_id:
            try:
                template = cls.get_template(db, int(template_id))
            except Exception:
                # Keep the create flow usable. If the selected template id is
                # stale/missing, we fall back to department + inventory_type.
                template = None

        # Then try to find a matching active system template.
        if not template and department and inv_type:
            template = (
                db.query(DynamicInventoryTemplate)
                .options(
                    joinedload(DynamicInventoryTemplate.template_fields),
                    joinedload(DynamicInventoryTemplate.template_calculation_rules),
                    joinedload(DynamicInventoryTemplate.template_metrics),
                )
                .filter(
                    DynamicInventoryTemplate.department == department,
                    DynamicInventoryTemplate.inventory_type == inv_type,
                    DynamicInventoryTemplate.is_active.is_(True),
                )
                .first()
            )

            if not template:
                cls.seed_default_templates(db, actor_user_id=actor_user_id)
                template = (
                    db.query(DynamicInventoryTemplate)
                    .options(
                        joinedload(DynamicInventoryTemplate.template_fields),
                        joinedload(DynamicInventoryTemplate.template_calculation_rules),
                        joinedload(DynamicInventoryTemplate.template_metrics),
                    )
                    .filter(
                        DynamicInventoryTemplate.department == department,
                        DynamicInventoryTemplate.inventory_type == inv_type,
                        DynamicInventoryTemplate.is_active.is_(True),
                    )
                    .first()
                )

        # If a full template exists, create the inventory from that template.
        if template:
            return cls.create_inventory(db, {
                "title": data.get("title") or template.name,
                "description": data.get("description") or template.description,
                "department": template.department,
                "inventory_type": template.inventory_type,
                "template_id": template.id,
                "report_title": data.get("report_title") or f"{data.get('title') or template.name} Report",
                "reporter_name": data.get("reporter_name"),
                "company_name": data.get("company_name"),
                "access_type": data.get("access_type", InventoryAccessType.ASSIGNED_USERS),
                "fields": [cls._field_from_template_payload(f) for f in template.template_fields],
                "calculation_rules": [cls._rule_from_template_payload(r) for r in template.template_calculation_rules],
                "metrics": [cls._metric_from_template_payload(m) for m in template.template_metrics],
                "user_access": [{"user_id": uid, "role": InventoryUserRole.EDITOR} for uid in data.get("assigned_user_ids", [])],
            }, actor_user_id=actor_user_id)

        # Final fallback: create a valid module even when a detailed predefined
        # template has not been built yet. This is important because the frontend
        # exposes every valid InventoryTemplateType enum, not only seeded ones.
        if not department:
            _http_error(422, "Department is required")
        if not inv_type:
            inv_type = InventoryTemplateType.CUSTOM

        title = data.get("title") or _enum_value(inv_type).replace("_", " ").title()
        return cls.create_inventory(db, {
            "title": title,
            "description": data.get("description") or f"General {title.lower()} records.",
            "department": department,
            "inventory_type": inv_type,
            "template_id": None,
            "report_title": data.get("report_title") or f"{title} Report",
            "reporter_name": data.get("reporter_name"),
            "company_name": data.get("company_name"),
            "access_type": data.get("access_type", InventoryAccessType.ASSIGNED_USERS),
            "fields": cls._generic_fields_for_inventory(department, inv_type),
            "calculation_rules": [],
            "metrics": [
                _metric("total_quantity", "Total Quantity", InventoryMetricType.SUM, field_key="quantity", order=1),
                _metric("total_amount", "Total Amount", InventoryMetricType.SUM, field_key="amount", unit="TZS", order=2),
            ],
            "user_access": [{"user_id": uid, "role": InventoryUserRole.EDITOR} for uid in data.get("assigned_user_ids", [])],
        }, actor_user_id=actor_user_id)

    @staticmethod
    def _field_from_template_payload(f: DynamicInventoryTemplateField) -> Dict[str, Any]:
        return {
            "field_name": f.field_name, "field_key": f.field_key, "field_code": f.field_code,
            "field_type": f.field_type, "field_direction": f.field_direction, "field_category": f.field_category,
            "lookup_group": f.lookup_group, "description": f.description, "placeholder": f.placeholder,
            "default_value": f.default_value, "unit_label": f.unit_label, "is_required": f.is_required,
            "is_unique": f.is_unique, "is_searchable": f.is_searchable, "is_filterable": f.is_filterable,
            "is_sortable": f.is_sortable, "is_indexed": f.is_indexed, "is_report_visible": f.is_report_visible,
            "is_dashboard_visible": f.is_dashboard_visible, "is_system_calculated": f.is_system_calculated,
            "is_user_editable": f.is_user_editable, "calculation_key": f.calculation_key,
            "options_json": f.options_json or [], "validation_json": f.validation_json or {},
            "display_json": f.display_json or {}, "settings_json": f.settings_json or {}, "order_index": f.order_index,
        }

    @staticmethod
    def _rule_from_template_payload(r: DynamicInventoryTemplateCalculationRule) -> Dict[str, Any]:
        return {
            "rule_key": r.rule_key, "rule_name": r.rule_name, "rule_type": r.rule_type, "scope": r.scope,
            "inputs_json": r.inputs_json or {}, "outputs_json": r.outputs_json or {},
            "expression_label": r.expression_label, "rule_config_json": r.rule_config_json or {},
            "is_active": r.is_active, "run_on_create": r.run_on_create, "run_on_update": r.run_on_update,
            "run_before_report": r.run_before_report, "order_index": r.order_index,
        }

    @staticmethod
    def _metric_from_template_payload(m: DynamicInventoryTemplateMetric) -> Dict[str, Any]:
        return {
            "metric_key": m.metric_key, "metric_type": m.metric_type, "label": m.label,
            "field_key": m.field_key, "calculation_rule_key": m.calculation_rule_key, "unit_label": m.unit_label,
            "is_enabled": m.is_enabled, "is_report_visible": m.is_report_visible, "is_dashboard_visible": m.is_dashboard_visible,
            "settings_json": m.settings_json or {}, "order_index": m.order_index,
        }

    @classmethod
    def create_inventory(cls, db: Session, payload: Any, actor_user_id: Optional[int] = None) -> DynamicInventory:
        data = _dump(payload)
        title = data.get("title")
        if not title:
            _http_error(422, "Inventory title is required")
        fields = data.pop("fields", []) or []
        rules = data.pop("calculation_rules", []) or []
        metrics = data.pop("metrics", []) or []
        user_access = data.pop("user_access", []) or []
        credentials = data.pop("credentials", []) or []
        department = _to_enum(RanchDepartment, data.get("department"))
        inventory = DynamicInventory(
            title=title,
            slug=data.get("slug") or cls._unique_slug(db, title),
            description=data.get("description"),
            department=department,
            inventory_type=_to_enum(InventoryTemplateType, data.get("inventory_type"), InventoryTemplateType.CUSTOM),
            template_id=data.get("template_id"),
            report_title=data.get("report_title") or f"{title} Report",
            reporter_name=data.get("reporter_name"),
            company_name=data.get("company_name"),
            logo_url=data.get("logo_url"),
            access_type=_to_enum(InventoryAccessType, data.get("access_type"), InventoryAccessType.ASSIGNED_USERS),
            status=_to_enum(InventoryStatus, data.get("status"), InventoryStatus.ACTIVE),
            is_active=data.get("is_active", True),
            report_date_field_id=data.get("report_date_field_id"),
            calculation_profile_json=data.get("calculation_profile_json") or {},
            settings_json=data.get("settings_json") or {},
            report_config_json=data.get("report_config_json") or {},
            dashboard_config_json=data.get("dashboard_config_json") or {},
            created_by_user_id=actor_user_id,
            updated_by_user_id=actor_user_id,
        )
        db.add(inventory)
        db.flush()
        if not fields:
            fields = cls._classification_fields_for_department(department, start_order=1)
        field_map = cls._create_inventory_fields(db, inventory, fields, actor_user_id)
        rule_map = cls._create_inventory_rules(db, inventory, rules, actor_user_id)
        cls._create_inventory_metrics(db, inventory, metrics, field_map, rule_map, actor_user_id)
        for access in user_access:
            cls.create_access(db, inventory.id, access, actor_user_id=actor_user_id, commit=False)
        for cred in credentials:
            cls.create_credential(db, inventory.id, cred, actor_user_id=actor_user_id, commit=False)
        cls._audit(db, action="create_inventory", actor_user_id=actor_user_id, inventory_id=inventory.id, description=f"Created inventory {inventory.title}")
        db.commit()
        return cls.get_inventory(db, inventory.id)

    @classmethod
    def _create_inventory_fields(cls, db: Session, inventory: DynamicInventory, fields: Sequence[Any], actor_user_id: Optional[int]) -> Dict[str, DynamicInventoryField]:
        result: Dict[str, DynamicInventoryField] = {}
        seen = set()
        for index, payload in enumerate(fields):
            data = _dump(payload)
            key = _normalize_key(data.get("field_key") or data.get("field_name"))
            if key in seen:
                continue
            seen.add(key)
            field_type = _field_type_for_storage(data.get("field_type"), key=data.get("field_key") or data.get("field_name"), category=data.get("field_category"), unit=data.get("unit_label"))
            direction = _to_enum(InventoryFieldDirection, data.get("field_direction"), InventoryFieldDirection.INPUT)
            is_system = bool(data.get("is_system_calculated", direction == InventoryFieldDirection.OUTPUT or _enum_value(data.get("field_type")).lower() in {"system_calculated", "system", "calculated", "output"}))
            field = DynamicInventoryField(
                inventory_id=inventory.id,
                field_name=data.get("field_name") or key.replace("_", " ").title(),
                field_key=key,
                field_code=(data.get("field_code") or _field_code(index)).upper(),
                field_type=field_type,
                field_direction=direction,
                field_category=_to_enum(InventoryFieldCategory, data.get("field_category"), InventoryFieldCategory.GENERAL),
                lookup_group=_to_enum(InventoryLookupGroup, data.get("lookup_group"), None),
                description=data.get("description"),
                placeholder=data.get("placeholder"),
                default_value=data.get("default_value"),
                unit_label=data.get("unit_label"),
                is_required=bool(data.get("is_required", False)),
                is_unique=bool(data.get("is_unique", False)),
                is_searchable=bool(data.get("is_searchable", False)),
                is_filterable=bool(data.get("is_filterable", False)),
                is_sortable=bool(data.get("is_sortable", False)),
                is_indexed=bool(data.get("is_indexed", False)),
                is_report_visible=bool(data.get("is_report_visible", True)),
                is_dashboard_visible=bool(data.get("is_dashboard_visible", False)),
                is_system_calculated=is_system,
                is_user_editable=bool(data.get("is_user_editable", not is_system)),
                calculation_key=data.get("calculation_key"),
                formula_expression=data.get("formula_expression"),
                options_json=_jsonable(data.get("options_json") or []),
                validation_json=_jsonable(data.get("validation_json") or {}),
                display_json=_jsonable(data.get("display_json") or {}),
                settings_json=_jsonable(data.get("settings_json") or {}),
                order_index=data.get("order_index", index + 1),
                created_by_user_id=actor_user_id,
                updated_by_user_id=actor_user_id,
            )
            db.add(field)
            db.flush()
            result[key] = field
        return result

    @classmethod
    def _create_inventory_rules(cls, db: Session, inventory: DynamicInventory, rules: Sequence[Any], actor_user_id: Optional[int]) -> Dict[str, DynamicInventoryCalculationRule]:
        result: Dict[str, DynamicInventoryCalculationRule] = {}
        for index, payload in enumerate(rules):
            data = _dump(payload)
            key = data.get("rule_key") or _normalize_key(data.get("rule_name") or f"rule_{index + 1}")
            rule = DynamicInventoryCalculationRule(
                inventory_id=inventory.id,
                template_id=inventory.template_id,
                rule_key=key,
                rule_name=data.get("rule_name") or key.replace("_", " ").title(),
                rule_type=_to_enum(InventoryCalculationRuleType, data.get("rule_type"), _calc_type("CUSTOM_SYSTEM_RULE")),
                scope=_to_enum(InventoryCalculationScope, data.get("scope"), InventoryCalculationScope.ROW),
                inputs_json=data.get("inputs_json") or {},
                outputs_json=data.get("outputs_json") or {},
                expression_label=data.get("expression_label"),
                rule_config_json=data.get("rule_config_json") or {},
                is_active=data.get("is_active", True),
                is_system_rule=data.get("is_system_rule", True),
                run_on_create=data.get("run_on_create", True),
                run_on_update=data.get("run_on_update", True),
                run_before_report=data.get("run_before_report", True),
                order_index=data.get("order_index", index + 1),
                created_by_user_id=actor_user_id,
                updated_by_user_id=actor_user_id,
            )
            db.add(rule)
            db.flush()
            result[key] = rule
        return result

    @classmethod
    def _create_inventory_metrics(cls, db: Session, inventory: DynamicInventory, metrics: Sequence[Any], field_map: Dict[str, DynamicInventoryField], rule_map: Dict[str, DynamicInventoryCalculationRule], actor_user_id: Optional[int]) -> Dict[str, DynamicInventoryMetric]:
        result = {}
        for index, payload in enumerate(metrics):
            data = _dump(payload)
            key = data.get("metric_key") or _normalize_key(data.get("label") or f"metric_{index + 1}")
            field_id = data.get("field_id")
            if not field_id and data.get("field_key"):
                field_id = getattr(field_map.get(data.get("field_key")), "id", None)
            rule_id = data.get("calculation_rule_id")
            if not rule_id and data.get("calculation_rule_key"):
                rule_id = getattr(rule_map.get(data.get("calculation_rule_key")), "id", None)
            metric = DynamicInventoryMetric(
                inventory_id=inventory.id,
                field_id=field_id,
                calculation_rule_id=rule_id,
                metric_key=key,
                metric_type=_to_enum(InventoryMetricType, data.get("metric_type"), InventoryMetricType.SUM),
                label=data.get("label") or key.replace("_", " ").title(),
                description=data.get("description"),
                unit_label=data.get("unit_label"),
                is_enabled=data.get("is_enabled", True),
                is_report_visible=data.get("is_report_visible", True),
                is_dashboard_visible=data.get("is_dashboard_visible", True),
                order_index=data.get("order_index", index + 1),
                formula_expression=data.get("formula_expression"),
                settings_json=data.get("settings_json") or {},
                created_by_user_id=actor_user_id,
                updated_by_user_id=actor_user_id,
            )
            db.add(metric)
            db.flush()
            result[key] = metric
        return result

    @classmethod
    def list_inventories(cls, db: Session, department: Optional[Union[str, RanchDepartment]] = None, inventory_type: Optional[Union[str, InventoryTemplateType]] = None, include_archived: bool = False, include_deleted: bool = False, user_id: Optional[int] = None, is_admin: bool = True, search: Optional[str] = None) -> List[DynamicInventory]:
        q = db.query(DynamicInventory).options(joinedload(DynamicInventory.fields), joinedload(DynamicInventory.periods).joinedload(DynamicInventoryPeriod.rows), joinedload(DynamicInventory.metrics), joinedload(DynamicInventory.alerts))
        if not include_deleted:
            q = q.filter(DynamicInventory.deleted_at.is_(None))
        if not include_archived:
            q = q.filter(DynamicInventory.status != InventoryStatus.ARCHIVED)
        if department:
            q = q.filter(DynamicInventory.department == _to_enum(RanchDepartment, department))
        if inventory_type:
            q = q.filter(DynamicInventory.inventory_type == _to_enum(InventoryTemplateType, inventory_type))
        if search:
            like = f"%{search}%"
            q = q.filter(or_(DynamicInventory.title.ilike(like), DynamicInventory.description.ilike(like)))
        if user_id and not is_admin:
            q = q.join(DynamicInventoryUserAccess).filter(DynamicInventoryUserAccess.user_id == user_id, DynamicInventoryUserAccess.is_active.is_(True))
        return q.order_by(DynamicInventory.department, DynamicInventory.title).all()

    @classmethod
    def list_inventory_summaries(cls, db: Session, **kwargs: Any) -> List[Dict[str, Any]]:
        return [cls._inventory_summary(inv) for inv in cls.list_inventories(db, **kwargs)]

    @classmethod
    def get_inventory(cls, db: Session, inventory_id: int, include_deleted: bool = False) -> DynamicInventory:
        q = db.query(DynamicInventory).options(joinedload(DynamicInventory.fields), joinedload(DynamicInventory.calculation_rules), joinedload(DynamicInventory.metrics), joinedload(DynamicInventory.user_access), joinedload(DynamicInventory.credentials), joinedload(DynamicInventory.periods).joinedload(DynamicInventoryPeriod.rows).joinedload(DynamicInventoryRow.values), joinedload(DynamicInventory.alerts)).filter(DynamicInventory.id == inventory_id)
        if not include_deleted:
            q = q.filter(DynamicInventory.deleted_at.is_(None))
        inventory = q.first()
        if not inventory:
            _http_error(404, "Inventory not found")
        return inventory

    @classmethod
    def update_inventory(cls, db: Session, inventory_id: int, payload: Any, actor_user_id: Optional[int] = None) -> DynamicInventory:
        inventory = cls.get_inventory(db, inventory_id)
        data = _dump(payload, exclude_unset=True)
        if "title" in data and data["title"]:
            inventory.title = data["title"]
            if not data.get("slug"):
                inventory.slug = cls._unique_slug(db, data["title"], inventory_id=inventory.id)
        if "slug" in data and data["slug"]:
            inventory.slug = cls._unique_slug(db, data["slug"], inventory_id=inventory.id)
        for attr in ["description", "report_title", "reporter_name", "company_name", "logo_url", "report_date_field_id"]:
            if attr in data:
                setattr(inventory, attr, data[attr])
        if "department" in data and data["department"] is not None:
            inventory.department = _to_enum(RanchDepartment, data["department"])
        if "inventory_type" in data and data["inventory_type"] is not None:
            inventory.inventory_type = _to_enum(InventoryTemplateType, data["inventory_type"])
        if "access_type" in data and data["access_type"] is not None:
            inventory.access_type = _to_enum(InventoryAccessType, data["access_type"])
        if "status" in data and data["status"] is not None:
            inventory.status = _to_enum(InventoryStatus, data["status"])
        if "is_active" in data:
            inventory.is_active = bool(data["is_active"])
        for json_attr in ["calculation_profile_json", "settings_json", "report_config_json", "dashboard_config_json"]:
            if json_attr in data and data[json_attr] is not None:
                setattr(inventory, json_attr, data[json_attr])
        inventory.updated_by_user_id = actor_user_id
        cls._audit(db, action="update_inventory", actor_user_id=actor_user_id, inventory_id=inventory.id, description=f"Updated inventory {inventory.title}")
        db.commit()
        return cls.get_inventory(db, inventory.id)

    @classmethod
    def archive_inventory(cls, db: Session, inventory_id: int, actor_user_id: Optional[int] = None) -> DynamicInventory:
        inventory = cls.get_inventory(db, inventory_id)
        inventory.status = InventoryStatus.ARCHIVED
        inventory.is_active = False
        inventory.archived_at = _now()
        inventory.archived_by_user_id = actor_user_id
        cls._audit(db, action="archive_inventory", actor_user_id=actor_user_id, inventory_id=inventory.id, description=f"Archived {inventory.title}")
        db.commit()
        return cls.get_inventory(db, inventory_id, include_deleted=True)

    @classmethod
    def restore_inventory(cls, db: Session, inventory_id: int, actor_user_id: Optional[int] = None) -> DynamicInventory:
        inventory = cls.get_inventory(db, inventory_id, include_deleted=True)
        inventory.status = InventoryStatus.ACTIVE
        inventory.is_active = True
        inventory.archived_at = None
        inventory.deleted_at = None
        inventory.archived_by_user_id = None
        inventory.deleted_by_user_id = None
        cls._audit(db, action="restore_inventory", actor_user_id=actor_user_id, inventory_id=inventory.id, description=f"Restored {inventory.title}")
        db.commit()
        return cls.get_inventory(db, inventory_id)

    @classmethod
    def delete_inventory(cls, db: Session, inventory_id: int, permanent: bool = False, actor_user_id: Optional[int] = None) -> Dict[str, Any]:
        inventory = cls.get_inventory(db, inventory_id, include_deleted=True)
        if permanent:
            db.delete(inventory)
            db.commit()
            return {"message": "Inventory permanently deleted", "id": inventory_id}
        inventory.status = InventoryStatus.DELETED
        inventory.is_active = False
        inventory.deleted_at = _now()
        inventory.deleted_by_user_id = actor_user_id
        cls._audit(db, action="delete_inventory", actor_user_id=actor_user_id, inventory_id=inventory.id, description=f"Deleted {inventory.title}")
        db.commit()
        return {"message": "Inventory deleted", "id": inventory_id}

    # ------------------------------------------------------------------
    # Periods
    # ------------------------------------------------------------------

    @staticmethod
    def _period_bounds(period_date: date, period_type: InventoryPeriodType) -> Tuple[date, date]:
        if period_type == InventoryPeriodType.WEEKLY:
            start = period_date - timedelta(days=period_date.weekday())
            return start, start + timedelta(days=6)
        if period_type == InventoryPeriodType.MONTHLY:
            start = date(period_date.year, period_date.month, 1)
            return start, date(period_date.year, period_date.month, monthrange(period_date.year, period_date.month)[1])
        if period_type == InventoryPeriodType.YEARLY:
            return date(period_date.year, 1, 1), date(period_date.year, 12, 31)
        return period_date, period_date

    @classmethod
    def get_or_create_period(cls, db: Session, inventory_id: int, period_date: Optional[date] = None, period_type: Union[str, InventoryPeriodType] = InventoryPeriodType.DAILY, actor_user_id: Optional[int] = None) -> DynamicInventoryPeriod:
        inventory = cls.get_inventory(db, inventory_id)
        period_date = period_date or date.today()
        period_type = _to_enum(InventoryPeriodType, period_type, InventoryPeriodType.DAILY)
        period = db.query(DynamicInventoryPeriod).filter(DynamicInventoryPeriod.inventory_id == inventory_id, DynamicInventoryPeriod.period_type == period_type, DynamicInventoryPeriod.period_date == period_date).first()
        if period:
            return period
        start_date, end_date = cls._period_bounds(period_date, period_type)
        period = DynamicInventoryPeriod(
            inventory_id=inventory_id,
            period_type=period_type,
            period_date=period_date,
            start_date=start_date,
            end_date=end_date,
            title=f"{inventory.title} - {period_date.isoformat()}",
            status=InventoryPeriodStatus.DRAFT,
            created_by_user_id=actor_user_id,
            updated_by_user_id=actor_user_id,
        )
        db.add(period)
        db.flush()
        cls._audit(db, action="create_period", actor_user_id=actor_user_id, inventory_id=inventory_id, period_id=period.id, description=f"Created period {period.title}")
        db.commit()
        db.refresh(period)
        return period

    @classmethod
    def get_or_create_today_period(cls, db: Session, inventory_id: int, actor_user_id: Optional[int] = None) -> Dict[str, Any]:
        period = cls.get_or_create_period(db, inventory_id, date.today(), InventoryPeriodType.DAILY, actor_user_id=actor_user_id)
        return cls.get_period_detail(db, period.id)

    @classmethod
    def create_period(cls, db: Session, inventory_id: int, payload: Any, actor_user_id: Optional[int] = None) -> DynamicInventoryPeriod:
        data = _dump(payload)
        return cls.get_or_create_period(db, inventory_id, _date_from_any(data.get("period_date")) or date.today(), _to_enum(InventoryPeriodType, data.get("period_type"), InventoryPeriodType.DAILY), actor_user_id=actor_user_id)

    @classmethod
    def get_period(cls, db: Session, period_id: int) -> DynamicInventoryPeriod:
        period = db.query(DynamicInventoryPeriod).options(joinedload(DynamicInventoryPeriod.rows).joinedload(DynamicInventoryRow.values).joinedload(DynamicInventoryValue.field)).filter(DynamicInventoryPeriod.id == period_id).first()
        if not period:
            _http_error(404, "Period not found")
        return period

    @classmethod
    def _period_needs_recalculation(cls, period: DynamicInventoryPeriod) -> bool:
        """Return True only when the period has no usable cached summary.

        This keeps normal reads and exports fast. Draft periods with a missing
        summary are recalculated; approved/submitted periods keep their cached
        approved totals unless explicit recalculation is called elsewhere.
        """
        summary = period.summary_json if isinstance(period.summary_json, dict) else {}
        if not summary:
            return True
        if "row_count" not in summary and "metrics" not in summary:
            return True
        return False

    @classmethod
    def get_period_detail(cls, db: Session, period_id: int) -> Dict[str, Any]:
        period = cls.get_period(db, period_id)
        inventory = cls.get_inventory(db, period.inventory_id)
        fields = [f for f in inventory.fields if f.deleted_at is None and f.is_active and not f.is_archived]

        # Performance: do not recalculate approved/locked/submitted periods on every GET.
        # Row create/update/submit/approve already refreshes summary_json. Recalculating
        # on every read makes detail/history/report/export endpoints slow.
        if cls._period_needs_recalculation(period):
            cls.recalculate_period(db, period.id, commit=True)
            db.refresh(period)

        summary = period.summary_json or cls.calculate_summary(db, inventory.id, period.start_date, period.end_date)
        return {
            **cls._period_out(period),
            "fields": fields,
            "rows": [cls._row_out(r, fields) for r in period.rows if not r.is_deleted],
            "summary": summary,
        }

    @classmethod
    def _ensure_period_editable(cls, period: DynamicInventoryPeriod) -> None:
        if period.status in {InventoryPeriodStatus.APPROVED, InventoryPeriodStatus.LOCKED, InventoryPeriodStatus.ARCHIVED} or period.locked_at:
            _http_error(409, "This period is approved/locked and cannot be edited.")

    @classmethod
    def submit_period(cls, db: Session, period_id: int, payload: Any = None, actor_user_id: Optional[int] = None) -> DynamicInventoryPeriod:
        period = cls.get_period(db, period_id)
        data = _dump(payload)
        cls.recalculate_period(db, period_id, commit=False)
        period.status = InventoryPeriodStatus.SUBMITTED
        period.notes = data.get("notes") or period.notes
        period.submitted_by_user_id = actor_user_id
        period.submitted_at = _now()
        period.updated_by_user_id = actor_user_id
        period.summary_json = cls.calculate_summary(db, period.inventory_id, period.start_date, period.end_date)
        cls._audit(db, action="submit_period", actor_user_id=actor_user_id, inventory_id=period.inventory_id, period_id=period.id, description=f"Submitted {period.title}")
        db.commit()
        db.refresh(period)
        return period

    @classmethod
    def approve_period(cls, db: Session, period_id: int, payload: Any = None, actor_user_id: Optional[int] = None) -> DynamicInventoryPeriod:
        period = cls.get_period(db, period_id)
        data = _dump(payload)
        cls.recalculate_period(db, period_id, commit=False)
        period.status = InventoryPeriodStatus.APPROVED
        period.notes = data.get("notes") or period.notes
        period.approved_by_user_id = actor_user_id
        period.approved_at = _now()
        period.locked_by_user_id = actor_user_id
        period.locked_at = period.approved_at
        period.updated_by_user_id = actor_user_id
        period.summary_json = cls.calculate_summary(db, period.inventory_id, period.start_date, period.end_date)
        cls._audit(db, action="approve_period", actor_user_id=actor_user_id, inventory_id=period.inventory_id, period_id=period.id, description=f"Approved {period.title}")
        db.commit()
        db.refresh(period)
        return period

    @classmethod
    def reject_period(cls, db: Session, period_id: int, payload: Any = None, actor_user_id: Optional[int] = None) -> DynamicInventoryPeriod:
        period = cls.get_period(db, period_id)
        data = _dump(payload)
        period.status = InventoryPeriodStatus.REJECTED
        period.rejection_reason = data.get("reason") or data.get("notes")
        period.rejected_by_user_id = actor_user_id
        period.rejected_at = _now()
        period.locked_at = None
        period.locked_by_user_id = None
        cls._audit(db, action="reject_period", actor_user_id=actor_user_id, inventory_id=period.inventory_id, period_id=period.id, description=f"Rejected {period.title}")
        db.commit()
        db.refresh(period)
        return period

    @classmethod
    def lock_period(cls, db: Session, period_id: int, actor_user_id: Optional[int] = None) -> DynamicInventoryPeriod:
        period = cls.get_period(db, period_id)
        period.status = InventoryPeriodStatus.LOCKED
        period.locked_by_user_id = actor_user_id
        period.locked_at = _now()
        db.commit()
        db.refresh(period)
        return period

    # ------------------------------------------------------------------
    # Rows, values, and calculation engine
    # ------------------------------------------------------------------

    @classmethod
    def create_row(cls, db: Session, inventory_id: int, period_id: int, payload: Any, actor_user_id: Optional[int] = None) -> Dict[str, Any]:
        inventory = cls.get_inventory(db, inventory_id)
        period = cls.get_period(db, period_id)
        if period.inventory_id != inventory_id:
            _http_error(400, "Period does not belong to this inventory")
        cls._ensure_period_editable(period)
        data = _dump(payload)
        values = data.get("values") or {}
        max_row = db.query(func.max(DynamicInventoryRow.row_number)).filter(DynamicInventoryRow.period_id == period_id).scalar() or 0
        row = DynamicInventoryRow(
            inventory_id=inventory_id,
            period_id=period_id,
            row_number=data.get("row_number") or int(max_row) + 1,
            row_label=data.get("row_label"),
            record_date=_date_from_any(values.get("date") or values.get("record_date") or data.get("record_date")) or period.period_date,
            metadata_json=data.get("metadata_json") or {},
            created_by_user_id=actor_user_id,
            updated_by_user_id=actor_user_id,
        )
        db.add(row)
        db.flush()
        fields = [f for f in inventory.fields if f.deleted_at is None and f.is_active and not f.is_archived]
        cls._save_input_values(db, inventory, row, fields, values, actor_user_id)
        # Make newly-created input DynamicInventoryValue rows visible to the
        # calculation engine before reading row.values. Without this flush,
        # relationship collections can still look empty and all outputs become 0.
        db.flush()
        cls._set_primary_entity(row, values)
        cls.calculate_row_outputs(db, inventory, row, fields, actor_user_id=actor_user_id)
        # Cache period summary immediately so subsequent detail/report/export calls
        # do not need to recalculate all rows.
        db.flush()
        period.summary_json = cls.calculate_summary(db, inventory_id, period.start_date, period.end_date)
        period.updated_by_user_id = actor_user_id
        cls._audit(db, action="create_row", actor_user_id=actor_user_id, inventory_id=inventory_id, period_id=period_id, row_id=row.id, description=f"Created row {row.row_number}")
        db.commit()
        row = db.query(DynamicInventoryRow).options(joinedload(DynamicInventoryRow.values).joinedload(DynamicInventoryValue.field)).filter(DynamicInventoryRow.id == row.id).first()
        return cls._row_out(row, fields)

    @classmethod
    def _save_input_values(cls, db: Session, inventory: DynamicInventory, row: DynamicInventoryRow, fields: Sequence[DynamicInventoryField], values: Dict[str, Any], actor_user_id: Optional[int]) -> None:
        by_key = {f.field_key: f for f in fields}
        by_id = {f.id: f for f in fields}
        by_code = {str(f.field_code).upper(): f for f in fields if f.field_code}
        saved = set()
        for raw_key, raw_value in values.items():
            key = str(raw_key)
            field = by_key.get(key) or by_key.get(_normalize_key(key)) or by_code.get(key.upper()) or (by_id.get(int(key)) if key.isdigit() else None)
            if not field or field.id in saved:
                continue
            if field.field_direction == InventoryFieldDirection.OUTPUT or field.is_system_calculated or not field.is_user_editable:
                continue
            cls._upsert_cell(db, inventory.id, row.id, field, raw_value, actor_user_id=actor_user_id, is_system_value=False)
            saved.add(field.id)
        for field in fields:
            if field.id not in saved and field.field_direction == InventoryFieldDirection.INPUT and field.default_value not in (None, ""):
                cls._upsert_cell(db, inventory.id, row.id, field, field.default_value, actor_user_id=actor_user_id, is_system_value=False)

    @staticmethod
    def _set_primary_entity(row: DynamicInventoryRow, values: Dict[str, Any]) -> None:
        """Derive display entity fields without requiring physical DB columns.

        Older versions of DynamicInventoryRow may not have primary_entity_* columns.
        Store the derived values in metadata_json as the durable compatibility layer,
        and also set real attributes when the model version supports them.
        """
        metadata = dict(getattr(row, "metadata_json", None) or {})

        def put(attr: str, value: Any) -> None:
            if value is None or value == "":
                return
            metadata[attr] = str(value)
            if hasattr(row, attr):
                setattr(row, attr, str(value))

        for key in ["animal_type", "crop_type", "machine_type", "feed_type", "fertilizer_type", "part_name"]:
            if values.get(key):
                put("primary_entity_type", values[key])
                break
        for key in ["breed", "crop_variety", "machine_name", "part_name", "field_or_plot"]:
            if values.get(key):
                put("primary_entity_name", values[key])
                break
        for key in ["asset_code", "part_code", "receipt_number", "tag_number"]:
            if values.get(key):
                put("primary_entity_code", values[key])
                break

        row.metadata_json = metadata

    @classmethod
    def update_row(cls, db: Session, row_id: int, payload: Any, actor_user_id: Optional[int] = None) -> Dict[str, Any]:
        row = db.query(DynamicInventoryRow).options(joinedload(DynamicInventoryRow.values).joinedload(DynamicInventoryValue.field)).filter(DynamicInventoryRow.id == row_id, DynamicInventoryRow.is_deleted.is_(False)).first()
        if not row:
            _http_error(404, "Row not found")
        period = cls.get_period(db, row.period_id)
        cls._ensure_period_editable(period)
        inventory = cls.get_inventory(db, row.inventory_id)
        data = _dump(payload, exclude_unset=True)
        values = data.get("values") or {}
        if "row_label" in data:
            row.row_label = data["row_label"]
        if values:
            fields = [f for f in inventory.fields if f.deleted_at is None and f.is_active and not f.is_archived]
            cls._save_input_values(db, inventory, row, fields, values, actor_user_id)
            db.flush()
            cls._set_primary_entity(row, {**cls._row_values(row, fields), **values})
            cls.calculate_row_outputs(db, inventory, row, fields, actor_user_id=actor_user_id)
        cls._audit(db, action="update_row", actor_user_id=actor_user_id, inventory_id=row.inventory_id, period_id=row.period_id, row_id=row.id, description=f"Updated row {row.row_number}")
        db.commit()
        row = db.query(DynamicInventoryRow).options(joinedload(DynamicInventoryRow.values).joinedload(DynamicInventoryValue.field)).filter(DynamicInventoryRow.id == row_id).first()
        return cls._row_out(row, inventory.fields)

    @classmethod
    def delete_row(cls, db: Session, row_id: int, payload: Any = None, actor_user_id: Optional[int] = None) -> Dict[str, Any]:
        row = db.query(DynamicInventoryRow).filter(DynamicInventoryRow.id == row_id, DynamicInventoryRow.is_deleted.is_(False)).first()
        if not row:
            _http_error(404, "Row not found")
        period = cls.get_period(db, row.period_id)
        cls._ensure_period_editable(period)
        data = _dump(payload)
        row.is_deleted = True
        row.deleted_at = _now()
        row.deleted_by_user_id = actor_user_id
        row.delete_reason = data.get("reason") or data.get("delete_reason")
        cls._audit(db, action="delete_row", actor_user_id=actor_user_id, inventory_id=row.inventory_id, period_id=row.period_id, row_id=row.id, description=f"Deleted row {row.row_number}")
        db.commit()
        return {"message": "Row deleted", "id": row_id}

    @classmethod
    def _upsert_cell(cls, db: Session, inventory_id: int, row_id: int, field: DynamicInventoryField, value: Any, *, actor_user_id: Optional[int] = None, is_system_value: bool = False, calculation_rule_id: Optional[int] = None) -> DynamicInventoryValue:
        cell = db.query(DynamicInventoryValue).filter(DynamicInventoryValue.row_id == row_id, DynamicInventoryValue.field_id == field.id).first()
        if not cell:
            cell = DynamicInventoryValue(inventory_id=inventory_id, row_id=row_id, field_id=field.id, created_by_user_id=actor_user_id)
            db.add(cell)
        cell.updated_by_user_id = actor_user_id
        cell.is_system_value = is_system_value
        cell.calculation_rule_id = calculation_rule_id
        cell.value_text = None
        cell.value_number = None
        cell.value_date = None
        cell.value_boolean = None
        cell.value_json = None
        if value is None or value == "":
            cell.display_value = None
            return cell
        field_type = field.field_type
        if field_type in {InventoryFieldType.NUMBER, InventoryFieldType.INTEGER, InventoryFieldType.DECIMAL, InventoryFieldType.CURRENCY, InventoryFieldType.PERCENTAGE, getattr(InventoryFieldType, "SYSTEM_CALCULATED", InventoryFieldType.NUMBER), getattr(InventoryFieldType, "FORMULA", InventoryFieldType.NUMBER)}:
            number = _decimal_or_none(value)
            if number is not None:
                cell.value_number = number
                cell.display_value = _format_decimal(number)
            else:
                cell.value_text = str(value)
                cell.display_value = str(value)
        elif field_type in {InventoryFieldType.DATE, InventoryFieldType.DATETIME}:
            parsed = _date_from_any(value)
            if parsed:
                cell.value_date = parsed
                cell.display_value = parsed.isoformat()
            else:
                cell.value_text = str(value)
                cell.display_value = str(value)
        elif field_type == InventoryFieldType.BOOLEAN:
            b = value if isinstance(value, bool) else str(value).lower() in {"true", "1", "yes", "y"}
            cell.value_boolean = b
            cell.display_value = "Yes" if b else "No"
        elif isinstance(value, (dict, list)) or field_type in {InventoryFieldType.MULTI_SELECT, InventoryFieldType.FILE, InventoryFieldType.IMAGE}:
            cell.value_json = _jsonable(value)
            cell.display_value = json.dumps(_jsonable(value), ensure_ascii=False)
        else:
            cell.value_text = str(value)
            cell.display_value = str(value)
        return cell

    @classmethod
    def calculate_row_outputs(cls, db: Session, inventory: DynamicInventory, row: DynamicInventoryRow, fields: Optional[Sequence[DynamicInventoryField]] = None, actor_user_id: Optional[int] = None) -> Dict[str, Any]:
        fields = list(fields or [f for f in inventory.fields if f.deleted_at is None and f.is_active and not f.is_archived])
        field_by_key = {f.field_key: f for f in fields}
        # Always flush and read cell values from the database, not only from
        # row.values. When a row is new, cells are often added by row_id and the
        # relationship collection may not refresh until after commit. Reading
        # directly prevents system calculations from seeing an empty row and
        # returning 0.0 for every output.
        db.flush()
        row_values = cls._row_raw_values_from_db(db, row.id, fields)
        outputs: Dict[str, Any] = {}
        for rule in sorted([r for r in inventory.calculation_rules if r.is_active and r.scope == InventoryCalculationScope.ROW], key=lambda r: r.order_index):
            result = cls._calculate_rule(rule.rule_type, row_values, getattr(rule, "rule_key", None))
            for output_alias, field_key in (rule.outputs_json or {}).items():
                value = result.get(output_alias)
                if value is None and output_alias != field_key:
                    value = result.get(field_key)
                if value is None:
                    continue
                field = field_by_key.get(field_key)
                if not field:
                    continue
                cls._upsert_cell(db, inventory.id, row.id, field, value, actor_user_id=actor_user_id, is_system_value=True, calculation_rule_id=rule.id)
                row_values[field_key] = value
                outputs[field_key] = value
        row.computed_json = _jsonable(outputs)
        return outputs

    @classmethod
    def _row_raw_values(cls, row: DynamicInventoryRow, fields: Sequence[DynamicInventoryField]) -> Dict[str, Any]:
        by_id = {f.id: f for f in fields}
        data = {}
        for cell in row.values:
            field = by_id.get(cell.field_id)
            if field:
                data[field.field_key] = cls._cell_value(cell)
        return data

    @classmethod
    def _row_raw_values_from_db(cls, db: Session, row_id: int, fields: Sequence[DynamicInventoryField]) -> Dict[str, Any]:
        """Read row values directly from the database for reliable calculations.

        Newly-created DynamicInventoryValue records can be present in the session
        but not yet loaded on row.values. This helper reads them by row_id after
        db.flush(), so system-owned calculations such as Current Balance, Total
        Additions, Net Movement and Purchase Value use the actual input cells.
        """
        by_id = {f.id: f for f in fields}
        data: Dict[str, Any] = {}
        cells = (
            db.query(DynamicInventoryValue)
            .filter(DynamicInventoryValue.row_id == row_id)
            .all()
        )
        for cell in cells:
            field = by_id.get(cell.field_id)
            if field:
                data[field.field_key] = cls._cell_value(cell)
        return data

    @classmethod
    def _calculate_rule(cls, rule_type: Any, values: Dict[str, Any], rule_key: Optional[str] = None) -> Dict[str, Any]:
        """Run system-owned ranch calculations.

        The user never writes formulas. Templates store a rule_type when the
        model supports it; otherwise they store CUSTOM_SYSTEM_RULE and preserve
        the rule_key. This method accepts either one.
        """
        n = lambda key: _num(values.get(key))
        out: Dict[str, Any] = {}
        rule_name = _enum_value(rule_type).lower()
        key_name = str(rule_key or "").lower()
        calc = key_name if rule_name in {"custom_system_rule", "inventorycalculationruletype.custom_system_rule"} and key_name else rule_name

        if calc == "animal_current_balance":
            out["current_balance"] = n("opening_balance") + n("born") + n("bought") + n("transferred_in") - n("sold") - n("died") - n("transferred_out") - n("culled")
        elif calc == "animal_total_additions":
            out["total_additions"] = n("born") + n("bought") + n("transferred_in")
        elif calc == "animal_total_reductions":
            out["total_reductions"] = n("sold") + n("died") + n("transferred_out") + n("culled")
        elif calc == "animal_net_movement":
            out["net_movement"] = n("born") + n("bought") + n("transferred_in") - n("sold") - n("died") - n("transferred_out") - n("culled")
        elif calc == "animal_mortality_rate":
            denom = n("opening_balance") + n("born") + n("bought") + n("transferred_in")
            out["mortality_rate"] = (_safe_divide(n("died"), denom) or Decimal("0")) * Decimal("100")
        elif calc == "animal_birth_rate":
            denom = n("opening_balance") + n("bought") + n("transferred_in")
            out["birth_rate"] = (_safe_divide(n("born"), denom) or Decimal("0")) * Decimal("100")
        elif calc == "animal_sales_value":
            out["sales_value"] = n("sold") * n("selling_price")
        elif calc == "animal_purchase_value":
            out["purchase_value"] = n("bought") * n("purchase_price")
        elif calc in {
            "feed_closing_stock",
            "crop_closing_stock",
            "seed_closing_stock",
            "fertilizer_closing_stock",
            "chemical_closing_stock",
            "spare_parts_closing_stock",
            "generic_closing_stock",
        }:
            out["closing_stock"] = n("opening_stock") + n("harvested") + n("purchased") + n("produced") + n("received") - n("used") - n("sold") - n("damaged") - n("issued")
        elif calc == "feed_total_cost":
            out["total_feed_cost"] = (n("purchased") or n("quantity")) * n("unit_cost")
        elif calc == "feed_days_remaining":
            avg = n("average_daily_use") or n("used")
            out["estimated_days_remaining"] = _safe_divide(n("closing_stock"), avg) if avg else None
        elif calc == "crop_yield_per_acre":
            out["yield_per_acre"] = _safe_divide(n("harvest_quantity"), n("acres_planted"))
        elif calc == "crop_total_production_cost":
            out["total_production_cost"] = n("labour_cost") + n("transport_cost") + n("packaging_cost") + n("other_cost")
        elif calc == "crop_estimated_sales_value":
            quantity = n("harvest_quantity") or n("sold") or n("quantity")
            price = n("selling_price") or n("unit_price")
            out["estimated_sales_value"] = quantity * price
            out["sales_value"] = quantity * price
        elif calc == "crop_profit_loss":
            sales = n("estimated_sales_value") or (n("harvest_quantity") * n("selling_price"))
            cost = n("total_production_cost") or (n("labour_cost") + n("transport_cost") + n("packaging_cost") + n("other_cost"))
            out["profit_loss"] = sales - cost
        elif calc == "crop_cost_per_unit":
            cost = n("total_production_cost") or (n("labour_cost") + n("transport_cost") + n("packaging_cost") + n("other_cost"))
            out["cost_per_unit"] = _safe_divide(cost, n("harvest_quantity"))
        elif calc == "machine_running_hours":
            out["running_hours"] = n("closing_hours") - n("opening_hours")
        elif calc == "fuel_total_cost":
            out["total_fuel_cost"] = n("fuel_litres") * n("price_per_litre")
        elif calc == "fuel_per_hour":
            hours = n("running_hours") or (n("closing_hours") - n("opening_hours"))
            out["fuel_per_hour"] = _safe_divide(n("fuel_litres"), hours)
        elif calc == "spare_parts_remaining_value":
            stock = n("closing_stock") or (n("opening_stock") + n("purchased") - n("used") - n("damaged"))
            out["remaining_stock_value"] = stock * n("unit_cost")
        elif calc in {"maintenance_total_cost", "repair_total_cost"}:
            out["total_maintenance_cost"] = n("labour_cost") + n("parts_cost") + n("other_cost")
        elif calc == "service_due_status":
            next_date = _date_from_any(values.get("next_service_date"))
            out["service_due_status"] = "Due" if next_date and next_date <= date.today() else ("Scheduled" if next_date else "Not set")
        return out

    @classmethod
    def recalculate_period(cls, db: Session, period_id: int, commit: bool = True) -> Dict[str, Any]:
        period = cls.get_period(db, period_id)
        inventory = cls.get_inventory(db, period.inventory_id)
        fields = [f for f in inventory.fields if f.deleted_at is None and f.is_active and not f.is_archived]
        count = 0
        for row in period.rows:
            if not row.is_deleted:
                cls.calculate_row_outputs(db, inventory, row, fields)
                count += 1
        period.summary_json = cls.calculate_summary(db, inventory.id, period.start_date, period.end_date)
        if commit:
            db.commit()
        return {"rows_recalculated": count, "summary": period.summary_json}

    # ------------------------------------------------------------------
    # Reports and dashboard
    # ------------------------------------------------------------------

    @classmethod
    def _rows_for_range(cls, db: Session, inventory_id: int, start_date: date, end_date: date) -> List[DynamicInventoryRow]:
        return db.query(DynamicInventoryRow).join(DynamicInventoryPeriod, DynamicInventoryRow.period_id == DynamicInventoryPeriod.id).options(joinedload(DynamicInventoryRow.values).joinedload(DynamicInventoryValue.field)).filter(DynamicInventoryRow.inventory_id == inventory_id, DynamicInventoryRow.is_deleted.is_(False), DynamicInventoryPeriod.start_date <= end_date, DynamicInventoryPeriod.end_date >= start_date).order_by(DynamicInventoryPeriod.period_date, DynamicInventoryRow.row_number).all()

    @classmethod
    def calculate_summary(cls, db: Session, inventory_id: int, start_date: date, end_date: date) -> Dict[str, Any]:
        inventory = cls.get_inventory(db, inventory_id)
        fields = [f for f in inventory.fields if f.deleted_at is None and f.is_active and not f.is_archived]
        rows = cls._rows_for_range(db, inventory_id, start_date, end_date)
        input_totals: Dict[str, Decimal] = {}
        output_totals: Dict[str, Decimal] = {}
        numeric_values: Dict[str, List[Decimal]] = {}
        field_by_id = {f.id: f for f in fields}
        for row in rows:
            for cell in row.values:
                field = field_by_id.get(cell.field_id)
                if not field or cell.value_number is None:
                    continue
                number = Decimal(cell.value_number)
                target = output_totals if field.field_direction == InventoryFieldDirection.OUTPUT or field.is_system_calculated else input_totals
                target[field.field_key] = target.get(field.field_key, Decimal("0")) + number
                numeric_values.setdefault(field.field_key, []).append(number)
        metric_values = []
        for metric in inventory.metrics:
            if not metric.is_enabled or not metric.field:
                continue
            values = numeric_values.get(metric.field.field_key, [])
            value: Any = None
            if metric.metric_type == InventoryMetricType.SUM:
                value = sum(values, Decimal("0"))
            elif metric.metric_type == InventoryMetricType.AVERAGE:
                value = sum(values, Decimal("0")) / len(values) if values else Decimal("0")
            elif metric.metric_type == InventoryMetricType.COUNT:
                value = len(values)
            elif metric.metric_type == InventoryMetricType.MIN:
                value = min(values) if values else None
            elif metric.metric_type == InventoryMetricType.MAX:
                value = max(values) if values else None
            metric_values.append({"metric_id": metric.id, "metric_key": metric.metric_key, "label": metric.label, "metric_type": metric.metric_type, "value": _jsonable(value), "unit_label": metric.unit_label, "field_key": metric.field.field_key})
        return {"row_count": len(rows), "input_totals": _jsonable(input_totals), "output_totals": _jsonable(output_totals), "metrics": metric_values, "alerts": {"open": db.query(DynamicInventoryAlert).filter(DynamicInventoryAlert.inventory_id == inventory_id, DynamicInventoryAlert.status == InventoryAlertStatus.OPEN).count()}}

    @classmethod
    def get_history(cls, db: Session, inventory_id: int, start_date: Optional[date] = None, end_date: Optional[date] = None) -> Dict[str, Any]:
        inventory = cls.get_inventory(db, inventory_id)
        q = db.query(DynamicInventoryPeriod).filter(DynamicInventoryPeriod.inventory_id == inventory_id)
        if start_date:
            q = q.filter(DynamicInventoryPeriod.end_date >= start_date)
        if end_date:
            q = q.filter(DynamicInventoryPeriod.start_date <= end_date)
        return {"inventory": cls._inventory_summary(inventory), "periods": [cls._period_out(p) for p in q.order_by(DynamicInventoryPeriod.period_date.desc()).all()]}

    @classmethod
    def _cached_or_calculated_summary(cls, db: Session, inventory_id: int, start_date: date, end_date: date) -> Dict[str, Any]:
        """Use cached period summary when possible; fall back to calculation.

        Exact daily/monthly period exports often hit one period. Its summary_json
        is already updated on create/submit/approve, so using it avoids scanning
        and recalculating all values during each report/export.
        """
        periods = db.query(DynamicInventoryPeriod).filter(
            DynamicInventoryPeriod.inventory_id == inventory_id,
            DynamicInventoryPeriod.start_date <= end_date,
            DynamicInventoryPeriod.end_date >= start_date,
        ).all()
        if len(periods) == 1:
            summary = periods[0].summary_json if isinstance(periods[0].summary_json, dict) else {}
            if summary and ("row_count" in summary or "metrics" in summary):
                return summary
        return cls.calculate_summary(db, inventory_id, start_date, end_date)

    @classmethod
    def get_report_data(cls, db: Session, inventory_id: int, request: Any, actor_user_id: Optional[int] = None) -> Dict[str, Any]:
        req = _dump(request)
        start_date = _date_from_any(req.get("start_date"))
        end_date = _date_from_any(req.get("end_date"))
        if not start_date or not end_date:
            _http_error(422, "start_date and end_date are required")
        inventory = cls.get_inventory(db, inventory_id)
        fields = [f for f in inventory.fields if f.deleted_at is None and f.is_active and not f.is_archived]
        rows = cls._rows_for_range(db, inventory_id, start_date, end_date)

        # Performance: exports should not recalculate every period by default.
        # Use cached summary_json for a single exact period, and only recalculate
        # when force_recalculate=true is explicitly requested.
        if bool(req.get("force_recalculate", False)):
            for period_id in {r.period_id for r in rows}:
                cls.recalculate_period(db, period_id, commit=False)
            db.flush()

        summary = cls._cached_or_calculated_summary(db, inventory_id, start_date, end_date)
        return {
            "inventory": cls._inventory_summary(inventory),
            "report_title": inventory.report_title or f"{inventory.title} Report",
            "report_type": _to_enum(InventoryReportType, req.get("report_type"), InventoryReportType.CUSTOM_RANGE),
            "start_date": start_date,
            "end_date": end_date,
            "fields": fields if req.get("visible_fields_only", True) else inventory.fields,
            "rows": [cls._row_out(row, fields) for row in rows],
            "summary": summary,
            "generated_at": _now(),
        }

    @classmethod
    def export_report_excel(cls, db: Session, inventory_id: int, request: Any, actor_user_id: Optional[int] = None):
        if Workbook is None or StreamingResponse is None:
            _http_error(500, "Excel export requires openpyxl")
        data = cls.get_report_data(db, inventory_id, request, actor_user_id=actor_user_id)
        wb = Workbook()
        ws = wb.active
        ws.title = "Raw Data"
        fields = data["fields"]
        ws.append([data["report_title"]])
        ws.append([f"Period: {data['start_date']} to {data['end_date']}"])
        ws.append([])
        ws.append([f.field_name for f in fields])
        for cell in ws[4]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="F1F5F9")
            cell.alignment = Alignment(horizontal="center")
        for row in data["rows"]:
            values = row.get("values") or {}
            ws.append([values.get(f.field_key) for f in fields])
        for idx, field in enumerate(fields, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = min(max(len(field.field_name) + 2, 14), 35)
        summary_ws = wb.create_sheet("Summary")
        summary_ws.append(["Metric", "Value", "Unit"])
        summary_ws.append(["Total Rows", data["summary"].get("row_count"), ""])
        for metric in data["summary"].get("metrics", []):
            summary_ws.append([metric.get("label"), metric.get("value"), metric.get("unit_label") or ""])
        out = io.BytesIO()
        wb.save(out)
        out.seek(0)
        filename = f"{_slugify(data['inventory']['title'])}-{data['start_date']}-{data['end_date']}.xlsx"
        return StreamingResponse(out, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    @classmethod
    def export_report_csv(cls, db: Session, inventory_id: int, request: Any, actor_user_id: Optional[int] = None):
        if StreamingResponse is None:
            _http_error(500, "Streaming response unavailable")
        data = cls.get_report_data(db, inventory_id, request, actor_user_id=actor_user_id)
        output = io.StringIO()
        writer = csv.writer(output)
        fields = data["fields"]
        writer.writerow([f.field_name for f in fields])
        for row in data["rows"]:
            values = row.get("values") or {}
            writer.writerow([values.get(f.field_key) for f in fields])
        output.seek(0)
        filename = f"{_slugify(data['inventory']['title'])}-{data['start_date']}-{data['end_date']}.csv"
        return StreamingResponse(iter([output.getvalue()]), media_type="text/csv", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    @classmethod
    def export_report_pdf(cls, db: Session, inventory_id: int, request: Any, actor_user_id: Optional[int] = None):
        if StreamingResponse is None:
            _http_error(500, "Streaming response unavailable")
        data = cls.get_report_data(db, inventory_id, request, actor_user_id=actor_user_id)
        if colors is None:
            return StreamingResponse(iter([json.dumps(_jsonable(data), indent=2).encode("utf-8")]), media_type="application/json")
        out = io.BytesIO()
        doc = SimpleDocTemplate(out, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20)
        styles = getSampleStyleSheet()
        story: List[Any] = [
            Paragraph(str(data["report_title"]), styles["Title"]),
            Paragraph(f"Inventory: {data['inventory']['title']}", styles["Normal"]),
            Paragraph(f"Period: {data['start_date']} to {data['end_date']}", styles["Normal"]),
            Spacer(1, 10),
            Paragraph("Summary", styles["Heading2"]),
        ]
        summary_rows = [["Metric", "Value", "Unit"], ["Total Rows", data["summary"].get("row_count"), ""]]
        for metric in data["summary"].get("metrics", []):
            summary_rows.append([metric.get("label"), metric.get("value"), metric.get("unit_label") or ""])
        table = Table(summary_rows, repeatRows=1)
        table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")), ("GRID", (0, 0), (-1, -1), 0.25, colors.grey), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold")]))
        story.append(table)
        story.append(Spacer(1, 12))
        fields = data["fields"][:12]
        record_rows = [[f.field_name for f in fields]]
        for row in data["rows"][:80]:
            values = row.get("values") or {}
            record_rows.append([values.get(f.field_key, "") for f in fields])
        records_table = Table(record_rows, repeatRows=1)
        records_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f1f5f9")), ("GRID", (0, 0), (-1, -1), 0.25, colors.grey), ("FONTSIZE", (0, 0), (-1, -1), 8)]))
        story.append(Paragraph("Records", styles["Heading2"]))
        story.append(records_table)
        story.append(Spacer(1, 14))
        story.append(Paragraph("Prepared By: ____________________        Approved By: ____________________", styles["Normal"]))
        doc.build(story)
        out.seek(0)
        filename = f"{_slugify(data['inventory']['title'])}-{data['start_date']}-{data['end_date']}.pdf"
        return StreamingResponse(out, media_type="application/pdf", headers={"Content-Disposition": f'attachment; filename="{filename}"'})

    @classmethod
    def get_dashboard(cls, db: Session, user_id: Optional[int] = None, is_admin: bool = True) -> Dict[str, Any]:
        inventories = cls.list_inventories(db, user_id=user_id, is_admin=is_admin)
        grouped: Dict[str, List[Dict[str, Any]]] = {"crops": [], "animals": [], "machinery": []}
        for inv in inventories:
            grouped.setdefault(inv.department.value, []).append(cls._inventory_summary(inv))
        departments = []
        for department in [RanchDepartment.CROPS, RanchDepartment.ANIMALS, RanchDepartment.MACHINERY]:
            items = grouped.get(department.value, [])
            departments.append({
                "department": department,
                "title": "Machineries & Maintenance" if department == RanchDepartment.MACHINERY else f"{department.value.title()} Department",
                "inventory_count": len(items),
                "pending_approvals": sum(i.get("pending_approvals", 0) for i in items),
                "open_alerts": sum(i.get("open_alerts", 0) for i in items),
                "inventories": items,
            })
        return {"departments": departments, "total_inventories": len(inventories), "generated_at": _now()}

    # ------------------------------------------------------------------
    # Access, credentials, alerts, attachments, audit
    # ------------------------------------------------------------------

    @classmethod
    def create_access(cls, db: Session, inventory_id: int, payload: Any, actor_user_id: Optional[int] = None, commit: bool = True) -> DynamicInventoryUserAccess:
        data = _dump(payload)
        obj = DynamicInventoryUserAccess(
            inventory_id=inventory_id,
            user_id=data.get("user_id"),
            role=_to_enum(InventoryUserRole, data.get("role"), InventoryUserRole.EDITOR),
            can_add_rows=data.get("can_add_rows", True),
            can_edit_rows=data.get("can_edit_rows", True),
            can_delete_rows=data.get("can_delete_rows", False),
            can_submit_periods=data.get("can_submit_periods", True),
            can_approve_periods=data.get("can_approve_periods", False),
            can_view_history=data.get("can_view_history", True),
            can_export_reports=data.get("can_export_reports", True),
            can_manage_fields=data.get("can_manage_fields", False),
            can_manage_users=data.get("can_manage_users", False),
            can_manage_templates=data.get("can_manage_templates", False),
            can_manage_lookups=data.get("can_manage_lookups", False),
            is_active=data.get("is_active", True),
            settings_json=data.get("settings_json") or {},
            created_by_user_id=actor_user_id,
            updated_by_user_id=actor_user_id,
        )
        db.add(obj)
        if commit:
            db.commit()
            db.refresh(obj)
        return obj

    @classmethod
    def create_credential(cls, db: Session, inventory_id: int, payload: Any, actor_user_id: Optional[int] = None, commit: bool = True) -> DynamicInventoryCredential:
        data = _dump(payload)
        password = data.get("password") or data.get("raw_password")
        if not password:
            _http_error(422, "password is required")
        obj = DynamicInventoryCredential(
            inventory_id=inventory_id,
            username=data.get("username"),
            password_hash=cls._hash_password(password),
            role=_to_enum(InventoryUserRole, data.get("role"), InventoryUserRole.EDITOR),
            must_change_password=data.get("must_change_password", True),
            is_active=data.get("is_active", True),
            permissions_json=data.get("permissions_json") or {},
            settings_json=data.get("settings_json") or {},
            created_by_user_id=actor_user_id,
            updated_by_user_id=actor_user_id,
        )
        db.add(obj)
        if commit:
            db.commit()
            db.refresh(obj)
        return obj

    @classmethod
    def list_lookup_options(cls, db: Session, group: Optional[Union[str, InventoryLookupGroup]] = None, department: Optional[Union[str, RanchDepartment]] = None, active_only: bool = True) -> List[DynamicInventoryLookupOption]:
        q = db.query(DynamicInventoryLookupOption)
        if group:
            q = q.filter(DynamicInventoryLookupOption.group == _to_enum(InventoryLookupGroup, group))
        if department:
            q = q.filter(or_(DynamicInventoryLookupOption.department == _to_enum(RanchDepartment, department), DynamicInventoryLookupOption.department.is_(None)))
        if active_only:
            q = q.filter(DynamicInventoryLookupOption.is_active.is_(True))
        return q.order_by(DynamicInventoryLookupOption.group, DynamicInventoryLookupOption.order_index, DynamicInventoryLookupOption.label).all()

    @classmethod
    def create_lookup_option(cls, db: Session, payload: Any, actor_user_id: Optional[int] = None) -> DynamicInventoryLookupOption:
        data = _dump(payload)
        obj = DynamicInventoryLookupOption(
            department=_to_enum(RanchDepartment, data.get("department"), None),
            group=_to_enum(InventoryLookupGroup, data.get("group")),
            label=data.get("label"),
            value=data.get("value") or _normalize_key(data.get("label")),
            parent_group=_to_enum(InventoryLookupGroup, data.get("parent_group"), None),
            parent_value=data.get("parent_value"),
            description=data.get("description"),
            metadata_json=data.get("metadata_json") or {},
            is_system_option=data.get("is_system_option", False),
            is_active=data.get("is_active", True),
            order_index=data.get("order_index", 0),
            created_by_user_id=actor_user_id,
            updated_by_user_id=actor_user_id,
        )
        db.add(obj)
        db.commit()
        db.refresh(obj)
        return obj

    @classmethod
    def create_alert(cls, db: Session, inventory_id: int, payload: Any, actor_user_id: Optional[int] = None) -> DynamicInventoryAlert:
        data = _dump(payload)
        obj = DynamicInventoryAlert(
            inventory_id=inventory_id,
            period_id=data.get("period_id"),
            row_id=data.get("row_id"),
            alert_key=data.get("alert_key") or _normalize_key(data.get("title")),
            title=data.get("title"),
            message=data.get("message"),
            level=_to_enum(InventoryAlertLevel, data.get("level"), InventoryAlertLevel.INFO),
            status=_to_enum(InventoryAlertStatus, data.get("status"), InventoryAlertStatus.OPEN),
            threshold_json=data.get("threshold_json") or {},
            context_json=data.get("context_json") or {},
            created_by_user_id=actor_user_id,
            updated_by_user_id=actor_user_id,
        )
        db.add(obj)
        db.commit()
        db.refresh(obj)
        return obj

    @classmethod
    def list_alerts(cls, db: Session, inventory_id: Optional[int] = None, status: Optional[Union[str, InventoryAlertStatus]] = None) -> List[DynamicInventoryAlert]:
        q = db.query(DynamicInventoryAlert)
        if inventory_id:
            q = q.filter(DynamicInventoryAlert.inventory_id == inventory_id)
        if status:
            q = q.filter(DynamicInventoryAlert.status == _to_enum(InventoryAlertStatus, status))
        return q.order_by(DynamicInventoryAlert.created_at.desc()).all()

    @classmethod
    def acknowledge_alert(cls, db: Session, alert_id: int, actor_user_id: Optional[int] = None) -> DynamicInventoryAlert:
        alert = db.query(DynamicInventoryAlert).filter(DynamicInventoryAlert.id == alert_id).first()
        if not alert:
            _http_error(404, "Alert not found")
        alert.status = InventoryAlertStatus.ACKNOWLEDGED
        alert.acknowledged_by_user_id = actor_user_id
        alert.acknowledged_at = _now()
        db.commit()
        db.refresh(alert)
        return alert

    @classmethod
    def resolve_alert(cls, db: Session, alert_id: int, actor_user_id: Optional[int] = None) -> DynamicInventoryAlert:
        alert = db.query(DynamicInventoryAlert).filter(DynamicInventoryAlert.id == alert_id).first()
        if not alert:
            _http_error(404, "Alert not found")
        alert.status = InventoryAlertStatus.RESOLVED
        alert.resolved_by_user_id = actor_user_id
        alert.resolved_at = _now()
        db.commit()
        db.refresh(alert)
        return alert

    @classmethod
    def create_attachment(cls, db: Session, inventory_id: int, payload: Any, actor_user_id: Optional[int] = None) -> DynamicInventoryAttachment:
        data = _dump(payload)
        obj = DynamicInventoryAttachment(
            inventory_id=inventory_id,
            period_id=data.get("period_id"),
            row_id=data.get("row_id"),
            field_id=data.get("field_id"),
            file_name=data.get("file_name"),
            file_url=data.get("file_url"),
            content_type=data.get("content_type"),
            file_size_bytes=data.get("file_size_bytes"),
            description=data.get("description"),
            metadata_json=data.get("metadata_json") or {},
            created_by_user_id=actor_user_id,
            updated_by_user_id=actor_user_id,
        )
        db.add(obj)
        db.commit()
        db.refresh(obj)
        return obj

    @classmethod
    def list_audit_logs(cls, db: Session, inventory_id: Optional[int] = None, limit: int = 100) -> List[DynamicInventoryAuditLog]:
        q = db.query(DynamicInventoryAuditLog)
        if inventory_id:
            q = q.filter(DynamicInventoryAuditLog.inventory_id == inventory_id)
        return q.order_by(DynamicInventoryAuditLog.created_at.desc()).limit(limit).all()


InventoryService = DynamicInventoryService
