from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal, ROUND_HALF_UP
from io import BytesIO
from typing import Any, Optional, Sequence

from fastapi import HTTPException, status
from sqlalchemy.orm import Session

from app.schemas.saas import (
    SaaSAnimalSummaryRow,
    SaaSClientSummaryRow,
    SaaSDateRangeReportRequest,
    SaaSDailyReportRequest,
    SaaSExportRequest,
    SaaSMonthlyReportRequest,
    SaaSReportRow,
    SaaSReportTotals,
    SaaSWeeklyReportRequest,
)
from app.services.saas_service import SaaSService


DECIMAL_ZERO_2 = Decimal("0.00")
DECIMAL_ZERO_3 = Decimal("0.000")
TWOPLACES = Decimal("0.01")
THREEPLACES = Decimal("0.001")

GRID_GREY = "D9D9D9"
HEADER_GREY = "E5E7EB"
PDF_GRID_GREY = "#CFCFCF"
PDF_HEADER_GREY = "#E5E7EB"


@dataclass
class ExportedFile:
    filename: str
    media_type: str
    content: bytes


def normalize_text(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    value = value.strip()
    return value or None


def quantize_2(value: Any) -> Decimal:
    if value is None or value == "":
        return DECIMAL_ZERO_2
    if isinstance(value, Decimal):
        return value.quantize(TWOPLACES, rounding=ROUND_HALF_UP)
    return Decimal(str(value)).quantize(TWOPLACES, rounding=ROUND_HALF_UP)


def quantize_3(value: Any) -> Decimal:
    if value is None or value == "":
        return DECIMAL_ZERO_3
    if isinstance(value, Decimal):
        return value.quantize(THREEPLACES, rounding=ROUND_HALF_UP)
    return Decimal(str(value)).quantize(THREEPLACES, rounding=ROUND_HALF_UP)


def format_date(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "day") and hasattr(value, "month") and hasattr(value, "year"):
        return f"{value.day}/{value.month}/{value.year}"
    return str(value)


def format_export_date(value: Optional[date]) -> str:
    return format_date(value or date.today())


def format_int(value: Any) -> str:
    try:
        return f"{int(value or 0):,}"
    except Exception:
        return "0"


def format_money_2(value: Any) -> str:
    return f"{quantize_2(value):,.2f}"


def format_money_3(value: Any) -> str:
    return f"{quantize_3(value):,.3f}"


def sanitize_filename(value: str) -> str:
    value = normalize_text(value) or "slaughter_services_report"
    value = re.sub(r"[^\w\-. ]+", "", value)
    value = re.sub(r"\s+", "_", value.strip())
    return value[:150] or "slaughter_services_report"


class SaaSExportService:
    EXCEL_MEDIA_TYPE = (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    PDF_MEDIA_TYPE = "application/pdf"

    def __init__(self, service: Optional[SaaSService] = None) -> None:
        self.service = service or SaaSService()
        self._openpyxl: Optional[dict[str, Any]] = None
        self._reportlab: Optional[dict[str, Any]] = None

    # =====================================================================
    # LAZY IMPORTS
    # =====================================================================
    def _get_openpyxl(self) -> dict[str, Any]:
        if self._openpyxl is not None:
            return self._openpyxl

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Excel export requires openpyxl. Install it with: pip install openpyxl",
            ) from exc

        self._openpyxl = {
            "Workbook": Workbook,
            "Alignment": Alignment,
            "Border": Border,
            "Font": Font,
            "PatternFill": PatternFill,
            "Side": Side,
        }
        return self._openpyxl

    def _get_reportlab(self) -> dict[str, Any]:
        if self._reportlab is not None:
            return self._reportlab

        try:
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_LEFT
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import (
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="PDF export requires reportlab. Install it with: pip install reportlab",
            ) from exc

        self._reportlab = {
            "colors": colors,
            "TA_LEFT": TA_LEFT,
            "A4": A4,
            "landscape": landscape,
            "ParagraphStyle": ParagraphStyle,
            "getSampleStyleSheet": getSampleStyleSheet,
            "mm": mm,
            "Paragraph": Paragraph,
            "SimpleDocTemplate": SimpleDocTemplate,
            "Spacer": Spacer,
            "Table": Table,
            "TableStyle": TableStyle,
        }
        return self._reportlab

    # =====================================================================
    # PUBLIC API
    # =====================================================================
    def export(self, db: Session, payload: SaaSExportRequest) -> ExportedFile:
        context = self._build_report_context(db, payload)

        if payload.export_format == "excel":
            content = self._build_excel(context)
            return ExportedFile(
                filename=f"{context['file_stem']}.xlsx",
                media_type=self.EXCEL_MEDIA_TYPE,
                content=content,
            )

        if payload.export_format == "pdf":
            content = self._build_pdf(context)
            return ExportedFile(
                filename=f"{context['file_stem']}.pdf",
                media_type=self.PDF_MEDIA_TYPE,
                content=content,
            )

        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Unsupported export format.",
        )

    # =====================================================================
    # REPORT CONTEXT
    # =====================================================================
    def _build_report_context(
        self,
        db: Session,
        payload: SaaSExportRequest,
    ) -> dict[str, Any]:
        scope = payload.scope
        report_title = (
            normalize_text(payload.report_title) or "UMG Slaughter Services Report"
        )
        prepared_on = payload.prepared_on or date.today()
        prepared_by_name = normalize_text(payload.prepared_by_name) or "System"
        organization_name = (
            normalize_text(payload.organization_name) or "Union Meat Group"
        )

        if scope == "daily":
            report_request = SaaSDailyReportRequest(
                report_date=payload.report_date,
                client_name=payload.client_name,
                animal_type=payload.animal_type,
                is_active=payload.is_active,
                include_rows=payload.include_rows,
                include_totals=payload.include_totals,
                include_client_summary=payload.include_client_summary,
                include_animal_summary=payload.include_animal_summary,
                report_title=report_title,
                prepared_by_name=prepared_by_name,
                prepared_on=prepared_on,
                organization_name=organization_name,
            )
            report = self.service.get_daily_report(db, report_request)
            file_stem = sanitize_filename(
                payload.file_name
                or f"UMG_Slaughter_Services_Daily_{report.report_date.isoformat()}"
            )

        elif scope == "weekly":
            report_request = SaaSWeeklyReportRequest(
                reference_date=payload.reference_date,
                week_starts_on=payload.week_starts_on,
                client_name=payload.client_name,
                animal_type=payload.animal_type,
                is_active=payload.is_active,
                include_rows=payload.include_rows,
                include_totals=payload.include_totals,
                include_client_summary=payload.include_client_summary,
                include_animal_summary=payload.include_animal_summary,
                report_title=report_title,
                prepared_by_name=prepared_by_name,
                prepared_on=prepared_on,
                organization_name=organization_name,
            )
            report = self.service.get_weekly_report(db, report_request)
            file_stem = sanitize_filename(
                payload.file_name
                or (
                    "UMG_Slaughter_Services_Weekly_"
                    f"{report.week_start_date.isoformat()}_"
                    f"{report.week_end_date.isoformat()}"
                )
            )

        elif scope == "monthly":
            report_request = SaaSMonthlyReportRequest(
                month=payload.month,
                year=payload.year,
                client_name=payload.client_name,
                animal_type=payload.animal_type,
                is_active=payload.is_active,
                include_rows=payload.include_rows,
                include_totals=payload.include_totals,
                include_client_summary=payload.include_client_summary,
                include_animal_summary=payload.include_animal_summary,
                report_title=report_title,
                prepared_by_name=prepared_by_name,
                prepared_on=prepared_on,
                organization_name=organization_name,
            )
            report = self.service.get_monthly_report(db, report_request)
            file_stem = sanitize_filename(
                payload.file_name
                or f"UMG_Slaughter_Services_Monthly_{payload.year}_{payload.month:02d}"
            )

        elif scope == "range":
            report_request = SaaSDateRangeReportRequest(
                start_date=payload.start_date,
                end_date=payload.end_date,
                client_name=payload.client_name,
                animal_type=payload.animal_type,
                is_active=payload.is_active,
                include_rows=payload.include_rows,
                include_totals=payload.include_totals,
                include_client_summary=payload.include_client_summary,
                include_animal_summary=payload.include_animal_summary,
                report_title=report_title,
                prepared_by_name=prepared_by_name,
                prepared_on=prepared_on,
                organization_name=organization_name,
            )
            report = self.service.get_date_range_report(db, report_request)
            file_stem = sanitize_filename(
                payload.file_name
                or (
                    "UMG_Slaughter_Services_Range_"
                    f"{report.start_date.isoformat()}_{report.end_date.isoformat()}"
                )
            )

        else:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Unsupported report scope.",
            )

        meta = getattr(report, "meta", None)

        return {
            "scope": scope,
            "title": normalize_text(getattr(meta, "report_title", None))
            or report_title,
            "organization_name": normalize_text(
                getattr(meta, "organization_name", None)
            )
            or organization_name,
            "scope_label": normalize_text(getattr(meta, "scope_label", None))
            or scope.title(),
            "prepared_by_name": normalize_text(
                getattr(meta, "prepared_by_name", None)
            )
            or prepared_by_name,
            "prepared_on": getattr(meta, "prepared_on", None) or prepared_on,
            "generated_at": datetime.now(timezone.utc).strftime(
                "%Y-%m-%d %H:%M:%S UTC"
            ),
            "rows": getattr(report, "rows", []) or [],
            "totals": getattr(report, "totals", None) or SaaSReportTotals(),
            "client_summary": getattr(report, "client_summary", []) or [],
            "animal_summary": getattr(report, "animal_summary", []) or [],
            "file_stem": file_stem,
        }

    # =====================================================================
    # EXCEL (increased font sizes)
    # =====================================================================
    def _build_excel(self, context: dict[str, Any]) -> bytes:
        xl = self._get_openpyxl()
        Workbook = xl["Workbook"]

        workbook = Workbook()
        scope = context["scope"]

        if scope == "daily":
            ws = workbook.active
            ws.title = "Daily Report"
            self._write_excel_daily_sheet(ws, context)
            self._apply_excel_footer(ws)

        elif scope == "weekly":
            ws = workbook.active
            ws.title = "Weekly Report"
            self._write_excel_weekly_sheet(ws, context)
            self._apply_excel_footer(ws)

        else:
            ws = workbook.active
            ws.title = "Report"
            self._write_excel_monthly_or_range_sheet(ws, context)
            self._apply_excel_footer(ws)

            if context["client_summary"]:
                client_ws = workbook.create_sheet("Client Summary")
                self._write_excel_client_summary_sheet(client_ws, context)
                self._apply_excel_footer(client_ws)

            if context["animal_summary"]:
                animal_ws = workbook.create_sheet("Animal Summary")
                self._write_excel_animal_summary_sheet(animal_ws, context)
                self._apply_excel_footer(animal_ws)

        return self._save_workbook(workbook)

    def _save_workbook(self, workbook: Any) -> bytes:
        stream = BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def _apply_excel_footer(self, ws: Any) -> None:
        try:
            ws.oddFooter.right.text = "Page &[Page] of &N"
        except Exception:
            pass

    def _apply_excel_page_setup(self, ws: Any, freeze_cell: str) -> None:
        ws.freeze_panes = freeze_cell
        ws.sheet_view.showGridLines = True
        ws.page_setup.orientation = "landscape"
        ws.page_margins.left = 0.2
        ws.page_margins.right = 0.2
        ws.page_margins.top = 0.35
        ws.page_margins.bottom = 0.35
        ws.print_options.horizontalCentered = False
        ws.print_options.verticalCentered = False

    def _excel_styles(self) -> dict[str, Any]:
        xl = self._get_openpyxl()
        Side = xl["Side"]
        Border = xl["Border"]
        Font = xl["Font"]
        PatternFill = xl["PatternFill"]

        thin = Side(style="thin", color=GRID_GREY)
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        return {
            "org_font": Font(name="Arial", size=14, bold=True),
            "title_font": Font(name="Arial", size=16, bold=True),
            "meta_label_font": Font(name="Arial", size=11, bold=True),
            "meta_value_font": Font(name="Arial", size=11),
            "header_font": Font(name="Arial", size=10, bold=True),
            "body_font": Font(name="Arial", size=10),
            "bold_font": Font(name="Arial", size=10, bold=True),
            "border": border,
            "header_fill": PatternFill(fill_type="solid", fgColor=HEADER_GREY),
        }

    def _excel_alignment(self, **kwargs: Any) -> Any:
        Alignment = self._get_openpyxl()["Alignment"]
        return Alignment(**kwargs)

    def _write_excel_report_header(
        self,
        ws: Any,
        context: dict[str, Any],
        end_column: int,
    ) -> int:
        styles = self._excel_styles()
        last_col = self._excel_column_letter(end_column)

        ws.merge_cells(f"A1:{last_col}1")
        ws["A1"] = context["organization_name"]
        ws["A1"].font = styles["org_font"]
        ws["A1"].alignment = self._excel_alignment(
            horizontal="left",
            vertical="center",
        )

        ws.merge_cells(f"A2:{last_col}2")
        ws["A2"] = context["title"]
        ws["A2"].font = styles["title_font"]
        ws["A2"].alignment = self._excel_alignment(
            horizontal="left",
            vertical="center",
        )

        meta_rows = [
            (
                "Prepared by",
                context["prepared_by_name"],
                "Date",
                format_export_date(context["prepared_on"]),
            ),
            (
                "Report scope",
                context["scope_label"],
                "Generated",
                context["generated_at"],
            ),
        ]

        start_row = 3
        for row_index, (label1, value1, label2, value2) in enumerate(
            meta_rows,
            start=start_row,
        ):
            ws.cell(row=row_index, column=1, value=label1).font = styles[
                "meta_label_font"
            ]
            ws.cell(row=row_index, column=2, value=value1).font = styles[
                "meta_value_font"
            ]
            ws.cell(row=row_index, column=2).alignment = self._excel_alignment(
                horizontal="left",
                vertical="center",
            )

            label2_col = max(4, end_column - 3)
            value2_col = label2_col + 1
            value2_end = end_column

            ws.cell(row=row_index, column=label2_col, value=label2).font = styles[
                "meta_label_font"
            ]
            ws.cell(row=row_index, column=value2_col, value=value2).font = styles[
                "meta_value_font"
            ]
            ws.cell(row=row_index, column=value2_col).alignment = self._excel_alignment(
                horizontal="left",
                vertical="center",
            )

            if value2_end > value2_col:
                ws.merge_cells(
                    start_row=row_index,
                    start_column=value2_col,
                    end_row=row_index,
                    end_column=value2_end,
                )

            for col in range(1, end_column + 1):
                cell = ws.cell(row=row_index, column=col)
                cell.border = styles["border"]
                cell.alignment = self._excel_alignment(
                    horizontal="left",
                    vertical="center",
                )

        for row in range(1, 5):
            for col in range(1, end_column + 1):
                ws.cell(row=row, column=col).border = styles["border"]

        return 6

    def _write_excel_table_header(
        self,
        ws: Any,
        row: int,
        headers: Sequence[str],
    ) -> None:
        styles = self._excel_styles()
        for col_idx, label in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=label)
            cell.font = styles["header_font"]
            cell.fill = styles["header_fill"]
            cell.border = styles["border"]
            cell.alignment = self._excel_alignment(
                horizontal="left",
                vertical="center",
                wrap_text=True,
            )

    def _write_excel_daily_sheet(self, ws: Any, context: dict[str, Any]) -> None:
        styles = self._excel_styles()
        headers = [
            "S.No",
            "Date",
            "Client",
            "Animal type",
            "Total Animals",
            "Unit Price service (per head) USD",
            "Total Service Revenue USD",
            "Unit Price offal USD",
            "Total Offal Revenue USD",
            "Total Combined Revenue USD",
        ]

        header_row = self._write_excel_report_header(ws, context, len(headers))
        self._write_excel_table_header(ws, header_row, headers)

        row_num = header_row + 1
        rows: Sequence[SaaSReportRow] = context["rows"] or []

        if rows:
            for index, item in enumerate(rows, start=1):
                values = [
                    index,
                    format_date(item.service_date),
                    item.client_name or "",
                    item.animal_type or "",
                    int(item.total_animals or 0),
                    float(quantize_3(item.unit_price_per_head_usd)),
                    float(quantize_2(item.total_revenue_usd)),
                    float(quantize_3(item.unit_price_offal_usd)),
                    float(quantize_2(item.total_offal_revenue_usd)),
                    float(quantize_2(item.total_combined_revenue_usd)),
                ]

                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    cell.font = styles["body_font"]
                    cell.border = styles["border"]
                    cell.alignment = self._excel_alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=True,
                    )

                    if col_idx in (1, 5):
                        cell.number_format = "#,##0"
                    elif col_idx in (6, 8):
                        cell.number_format = "#,##0.000"
                    elif col_idx in (7, 9, 10):
                        cell.number_format = "#,##0.00"

                row_num += 1
        else:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)
            cell = ws.cell(
                row=row_num,
                column=1,
                value="No slaughter services found for this report.",
            )
            cell.font = styles["body_font"]
            cell.border = styles["border"]
            cell.alignment = self._excel_alignment(horizontal="left", vertical="center")
            row_num += 1

        totals = context["totals"]

        service_row = row_num
        ws.merge_cells(start_row=service_row, start_column=1, end_row=service_row, end_column=6)
        ws.cell(row=service_row, column=1, value="Total Service Rev").font = styles["bold_font"]
        ws.cell(row=service_row, column=1).alignment = self._excel_alignment(horizontal="left", vertical="center")
        ws.cell(
            row=service_row,
            column=7,
            value=float(quantize_2(totals.total_service_revenue_usd)),
        ).font = styles["bold_font"]
        ws.cell(row=service_row, column=7).number_format = "#,##0.00"

        offal_row = row_num + 1
        ws.merge_cells(start_row=offal_row, start_column=1, end_row=offal_row, end_column=8)
        ws.cell(row=offal_row, column=1, value="Total Offal Rev").font = styles["bold_font"]
        ws.cell(row=offal_row, column=1).alignment = self._excel_alignment(horizontal="left", vertical="center")
        ws.cell(
            row=offal_row,
            column=9,
            value=float(quantize_2(totals.total_offal_revenue_usd)),
        ).font = styles["bold_font"]
        ws.cell(row=offal_row, column=9).number_format = "#,##0.00"

        combined_row = row_num + 2
        ws.merge_cells(
            start_row=combined_row,
            start_column=1,
            end_row=combined_row,
            end_column=9,
        )
        ws.cell(row=combined_row, column=1, value="Total Combined Rev").font = styles["bold_font"]
        ws.cell(row=combined_row, column=1).alignment = self._excel_alignment(horizontal="left", vertical="center")
        ws.cell(
            row=combined_row,
            column=10,
            value=float(quantize_2(totals.total_combined_revenue_usd)),
        ).font = styles["bold_font"]
        ws.cell(row=combined_row, column=10).number_format = "#,##0.00"

        for r in (service_row, offal_row, combined_row):
            for c in range(1, 11):
                ws.cell(row=r, column=c).border = styles["border"]
                ws.cell(row=r, column=c).alignment = self._excel_alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True,
                )

        widths = {
            "A": 7,
            "B": 11,
            "C": 14,
            "D": 13,
            "E": 10,
            "F": 16,
            "G": 15,
            "H": 13,
            "I": 15,
            "J": 17,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        self._apply_excel_page_setup(ws, f"A{header_row + 1}")

    def _write_excel_weekly_sheet(self, ws: Any, context: dict[str, Any]) -> None:
        styles = self._excel_styles()
        headers = [
            "S.No",
            "Date",
            "Client",
            "Animal type",
            "Animals",
            "Unit/head USD",
            "Service Rev USD",
            "Unit offal USD",
            "Offal Rev USD",
            "Combined Rev USD",
        ]

        self._write_excel_report_header(ws, context, 10)

        summary_headers = ["Metric", "Value"]
        self._write_excel_table_header(ws, 6, summary_headers)

        totals = context["totals"]
        summary_rows = [
            ("Total Clients Served", int(totals.total_clients_served or 0)),
            ("Total Animals", int(totals.total_animals or 0)),
            (
                "Total Service Revenue USD",
                float(quantize_2(totals.total_service_revenue_usd)),
            ),
            (
                "Total Offal Revenue USD",
                float(quantize_2(totals.total_offal_revenue_usd)),
            ),
            (
                "Total Combined Revenue USD",
                float(quantize_2(totals.total_combined_revenue_usd)),
            ),
        ]

        summary_row_num = 7
        for label, value in summary_rows:
            ws.cell(row=summary_row_num, column=1, value=label).font = styles["body_font"]
            ws.cell(row=summary_row_num, column=2, value=value).font = styles["body_font"]
            ws.cell(row=summary_row_num, column=1).border = styles["border"]
            ws.cell(row=summary_row_num, column=2).border = styles["border"]
            ws.cell(row=summary_row_num, column=1).alignment = self._excel_alignment(horizontal="left", vertical="center")
            ws.cell(row=summary_row_num, column=2).alignment = self._excel_alignment(horizontal="left", vertical="center")
            if "Revenue" in label:
                ws.cell(row=summary_row_num, column=2).number_format = "#,##0.00"
            summary_row_num += 1

        detail_title_row = 13
        ws.merge_cells(
            start_row=detail_title_row,
            start_column=1,
            end_row=detail_title_row,
            end_column=10,
        )
        ws.cell(row=detail_title_row, column=1, value="Weekly Detailed Rows").font = styles["bold_font"]
        ws.cell(row=detail_title_row, column=1).alignment = self._excel_alignment(horizontal="left", vertical="center")

        header_row = 15
        self._write_excel_table_header(ws, header_row, headers)

        row_num = header_row + 1
        rows: Sequence[SaaSReportRow] = context["rows"] or []

        if rows:
            for index, item in enumerate(rows, start=1):
                values = [
                    index,
                    format_date(item.service_date),
                    item.client_name or "",
                    item.animal_type or "",
                    int(item.total_animals or 0),
                    float(quantize_3(item.unit_price_per_head_usd)),
                    float(quantize_2(item.total_revenue_usd)),
                    float(quantize_3(item.unit_price_offal_usd)),
                    float(quantize_2(item.total_offal_revenue_usd)),
                    float(quantize_2(item.total_combined_revenue_usd)),
                ]

                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    cell.font = styles["body_font"]
                    cell.border = styles["border"]
                    cell.alignment = self._excel_alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=True,
                    )

                    if col_idx in (1, 5):
                        cell.number_format = "#,##0"
                    elif col_idx in (6, 8):
                        cell.number_format = "#,##0.000"
                    elif col_idx in (7, 9, 10):
                        cell.number_format = "#,##0.00"

                row_num += 1
        else:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)
            cell = ws.cell(
                row=row_num,
                column=1,
                value="No slaughter services found for this report.",
            )
            cell.font = styles["body_font"]
            cell.border = styles["border"]
            cell.alignment = self._excel_alignment(horizontal="left", vertical="center")

        widths = {
            "A": 7,
            "B": 11,
            "C": 14,
            "D": 13,
            "E": 10,
            "F": 14,
            "G": 14,
            "H": 13,
            "I": 14,
            "J": 15,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        self._apply_excel_page_setup(ws, "A16")

    def _write_excel_monthly_or_range_sheet(
        self,
        ws: Any,
        context: dict[str, Any],
    ) -> None:
        styles = self._excel_styles()
        headers = [
            "S.No",
            "Date",
            "Client",
            "Animal type",
            "Animals",
            "Unit/head USD",
            "Service Rev USD",
            "Unit offal USD",
            "Offal Rev USD",
            "Combined Rev USD",
        ]

        header_row = self._write_excel_report_header(ws, context, len(headers))
        self._write_excel_table_header(ws, header_row, headers)

        row_num = header_row + 1
        rows: Sequence[SaaSReportRow] = context["rows"] or []

        if rows:
            for index, item in enumerate(rows, start=1):
                values = [
                    index,
                    format_date(item.service_date),
                    item.client_name or "",
                    item.animal_type or "",
                    int(item.total_animals or 0),
                    float(quantize_3(item.unit_price_per_head_usd)),
                    float(quantize_2(item.total_revenue_usd)),
                    float(quantize_3(item.unit_price_offal_usd)),
                    float(quantize_2(item.total_offal_revenue_usd)),
                    float(quantize_2(item.total_combined_revenue_usd)),
                ]

                for col_idx, value in enumerate(values, start=1):
                    cell = ws.cell(row=row_num, column=col_idx, value=value)
                    cell.font = styles["body_font"]
                    cell.border = styles["border"]
                    cell.alignment = self._excel_alignment(
                        horizontal="left",
                        vertical="center",
                        wrap_text=True,
                    )

                    if col_idx in (1, 5):
                        cell.number_format = "#,##0"
                    elif col_idx in (6, 8):
                        cell.number_format = "#,##0.000"
                    elif col_idx in (7, 9, 10):
                        cell.number_format = "#,##0.00"

                row_num += 1
        else:
            ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=10)
            cell = ws.cell(
                row=row_num,
                column=1,
                value="No slaughter services found for this report.",
            )
            cell.font = styles["body_font"]
            cell.border = styles["border"]
            cell.alignment = self._excel_alignment(horizontal="left", vertical="center")
            row_num += 1

        row_num += 1
        self._write_excel_table_header(ws, row_num, ["Metric", "Value"])
        row_num += 1

        totals = context["totals"]
        totals_rows = [
            ("Total Clients Served", int(totals.total_clients_served or 0)),
            ("Total Animals", int(totals.total_animals or 0)),
            (
                "Total Service Revenue USD",
                float(quantize_2(totals.total_service_revenue_usd)),
            ),
            (
                "Total Offal Revenue USD",
                float(quantize_2(totals.total_offal_revenue_usd)),
            ),
            (
                "Total Combined Revenue USD",
                float(quantize_2(totals.total_combined_revenue_usd)),
            ),
        ]

        for label, value in totals_rows:
            ws.cell(row=row_num, column=1, value=label).font = styles["body_font"]
            ws.cell(row=row_num, column=2, value=value).font = styles["body_font"]
            ws.cell(row=row_num, column=1).border = styles["border"]
            ws.cell(row=row_num, column=2).border = styles["border"]
            ws.cell(row=row_num, column=1).alignment = self._excel_alignment(horizontal="left", vertical="center")
            ws.cell(row=row_num, column=2).alignment = self._excel_alignment(horizontal="left", vertical="center")
            if "Revenue" in label:
                ws.cell(row=row_num, column=2).number_format = "#,##0.00"
            row_num += 1

        widths = {
            "A": 7,
            "B": 11,
            "C": 14,
            "D": 13,
            "E": 10,
            "F": 14,
            "G": 14,
            "H": 13,
            "I": 14,
            "J": 15,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        self._apply_excel_page_setup(ws, f"A{header_row + 1}")

    def _write_excel_client_summary_sheet(
        self,
        ws: Any,
        context: dict[str, Any],
    ) -> None:
        styles = self._excel_styles()
        headers = [
            "Client",
            "Rows",
            "Animals",
            "Service Revenue USD",
            "Offal Revenue USD",
            "Combined Revenue USD",
        ]

        self._write_excel_report_header(ws, context, len(headers))

        ws.merge_cells("A5:F5")
        ws["A5"] = "Client Summary"
        ws["A5"].font = styles["bold_font"]
        ws["A5"].alignment = self._excel_alignment(horizontal="left", vertical="center")

        header_row = 7
        self._write_excel_table_header(ws, header_row, headers)

        row_num = 8
        for item in context["client_summary"]:
            values = [
                item.client_name or "",
                int(item.rows_count or 0),
                int(item.total_animals or 0),
                float(quantize_2(item.total_service_revenue_usd)),
                float(quantize_2(item.total_offal_revenue_usd)),
                float(quantize_2(item.total_combined_revenue_usd)),
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = styles["body_font"]
                cell.border = styles["border"]
                cell.alignment = self._excel_alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True,
                )
                if col_idx in (4, 5, 6):
                    cell.number_format = "#,##0.00"
            row_num += 1

        widths = {"A": 18, "B": 10, "C": 10, "D": 16, "E": 16, "F": 18}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        self._apply_excel_page_setup(ws, "A8")

    def _write_excel_animal_summary_sheet(
        self,
        ws: Any,
        context: dict[str, Any],
    ) -> None:
        styles = self._excel_styles()
        headers = [
            "Animal Type",
            "Rows",
            "Animals",
            "Service Revenue USD",
            "Offal Revenue USD",
            "Combined Revenue USD",
        ]

        self._write_excel_report_header(ws, context, len(headers))

        ws.merge_cells("A5:F5")
        ws["A5"] = "Animal Summary"
        ws["A5"].font = styles["bold_font"]
        ws["A5"].alignment = self._excel_alignment(horizontal="left", vertical="center")

        header_row = 7
        self._write_excel_table_header(ws, header_row, headers)

        row_num = 8
        for item in context["animal_summary"]:
            values = [
                item.animal_type or "",
                int(item.rows_count or 0),
                int(item.total_animals or 0),
                float(quantize_2(item.total_service_revenue_usd)),
                float(quantize_2(item.total_offal_revenue_usd)),
                float(quantize_2(item.total_combined_revenue_usd)),
            ]

            for col_idx, value in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = styles["body_font"]
                cell.border = styles["border"]
                cell.alignment = self._excel_alignment(
                    horizontal="left",
                    vertical="center",
                    wrap_text=True,
                )
                if col_idx in (4, 5, 6):
                    cell.number_format = "#,##0.00"
            row_num += 1

        widths = {"A": 16, "B": 10, "C": 10, "D": 16, "E": 16, "F": 18}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        self._apply_excel_page_setup(ws, "A8")

    def _excel_column_letter(self, index: int) -> str:
        result = ""
        value = index
        while value > 0:
            value, remainder = divmod(value - 1, 26)
            result = chr(65 + remainder) + result
        return result

    # =====================================================================
    # PDF (Times New Roman, larger fonts, landscape)
    # =====================================================================
    def _build_pdf(self, context: dict[str, Any]) -> bytes:
        rl = self._get_reportlab()
        SimpleDocTemplate = rl["SimpleDocTemplate"]
        landscape = rl["landscape"]
        A4 = rl["A4"]
        mm = rl["mm"]
        Paragraph = rl["Paragraph"]
        Spacer = rl["Spacer"]

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            leftMargin=10 * mm,
            rightMargin=10 * mm,
            topMargin=10 * mm,
            bottomMargin=12 * mm,
        )

        styles = self._pdf_styles()
        scope = context["scope"]

        story: list[Any] = []
        story.extend(self._build_pdf_header(context, styles))
        story.append(Spacer(1, 4))

        if scope == "daily":
            story.append(
                self._build_pdf_daily_table(
                    context["rows"] or [],
                    context["totals"],
                    styles,
                )
            )
        elif scope == "weekly":
            story.append(Paragraph("Weekly Summary", styles["section"]))
            story.append(self._build_pdf_totals_table(context["totals"], styles))
            story.append(Spacer(1, 4))
            story.append(Paragraph("Weekly Detailed Rows", styles["section"]))
            story.append(self._build_pdf_rows_table(context["rows"] or [], styles))
        else:
            story.append(Paragraph("Detailed Rows", styles["section"]))
            story.append(self._build_pdf_rows_table(context["rows"] or [], styles))
            story.append(Spacer(1, 4))
            story.append(Paragraph("Totals", styles["section"]))
            story.append(self._build_pdf_totals_table(context["totals"], styles))

            if context["client_summary"]:
                story.append(Spacer(1, 4))
                story.append(Paragraph("Client Summary", styles["section"]))
                story.append(
                    self._build_pdf_client_summary_table(
                        context["client_summary"],
                        styles,
                    )
                )

            if context["animal_summary"]:
                story.append(Spacer(1, 4))
                story.append(Paragraph("Animal Summary", styles["section"]))
                story.append(
                    self._build_pdf_animal_summary_table(
                        context["animal_summary"],
                        styles,
                    )
                )

        doc.build(
            story,
            onFirstPage=self._draw_pdf_page_number,
            onLaterPages=self._draw_pdf_page_number,
        )
        return buffer.getvalue()

    def _pdf_styles(self) -> dict[str, Any]:
        rl = self._get_reportlab()
        getSampleStyleSheet = rl["getSampleStyleSheet"]
        ParagraphStyle = rl["ParagraphStyle"]
        TA_LEFT = rl["TA_LEFT"]

        sample = getSampleStyleSheet()
        return {
            "org": ParagraphStyle(
                "SaaSOrg",
                parent=sample["Normal"],
                fontName="Times-Bold",
                fontSize=14,
                leading=16,
                alignment=TA_LEFT,
                spaceAfter=2,
            ),
            "title": ParagraphStyle(
                "SaaSTitle",
                parent=sample["Title"],
                fontName="Times-Bold",
                fontSize=18,
                leading=22,
                alignment=TA_LEFT,
                spaceAfter=6,
            ),
            "meta": ParagraphStyle(
                "SaaSMeta",
                parent=sample["Normal"],
                fontName="Times-Roman",
                fontSize=10,
                leading=12,
                alignment=TA_LEFT,
                spaceAfter=1,
            ),
            "section": ParagraphStyle(
                "SaaSSection",
                parent=sample["Heading3"],
                fontName="Times-Bold",
                fontSize=12,
                leading=14,
                alignment=TA_LEFT,
                spaceBefore=4,
                spaceAfter=6,
            ),
            "th": ParagraphStyle(
                "SaaSTableHeader",
                parent=sample["Normal"],
                fontName="Times-Bold",
                fontSize=9,
                leading=10.5,
                alignment=TA_LEFT,
            ),
            "td": ParagraphStyle(
                "SaaSTableBody",
                parent=sample["Normal"],
                fontName="Times-Roman",
                fontSize=10,
                leading=11.5,
                alignment=TA_LEFT,
            ),
            "td_bold": ParagraphStyle(
                "SaaSTableBodyBold",
                parent=sample["Normal"],
                fontName="Times-Bold",
                fontSize=10,
                leading=11.5,
                alignment=TA_LEFT,
            ),
        }

    def _build_pdf_header(
        self,
        context: dict[str, Any],
        styles: dict[str, Any],
    ) -> list[Any]:
        Paragraph = self._get_reportlab()["Paragraph"]
        return [
            Paragraph(context["organization_name"], styles["org"]),
            Paragraph(context["title"], styles["title"]),
            Paragraph(f"Prepared by: {context['prepared_by_name']}", styles["meta"]),
            Paragraph(
                f"Date: {format_export_date(context['prepared_on'])}",
                styles["meta"],
            ),
            Paragraph(f"Report scope: {context['scope_label']}", styles["meta"]),
            Paragraph(f"Generated: {context['generated_at']}", styles["meta"]),
        ]

    def _draw_pdf_page_number(self, canvas, doc) -> None:
        mm = self._get_reportlab()["mm"]
        canvas.saveState()
        canvas.setFont("Times-Roman", 9)
        canvas.drawRightString(
            doc.pagesize[0] - 10 * mm,
            8 * mm,
            f"Page {canvas.getPageNumber()}",
        )
        canvas.restoreState()

    def _pdf_p(self, value: Any, style: Any) -> Any:
        Paragraph = self._get_reportlab()["Paragraph"]
        text = "" if value is None else str(value)
        safe_text = (
            text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )
        return Paragraph(safe_text, style)

    def _base_pdf_table_style(self) -> list[tuple[Any, ...]]:
        colors = self._get_reportlab()["colors"]
        return [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(PDF_HEADER_GREY)),
            ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(PDF_GRID_GREY)),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]

    def _build_pdf_daily_table(
        self,
        rows: Sequence[SaaSReportRow],
        totals: SaaSReportTotals,
        styles: dict[str, Any],
    ) -> Any:
        rl = self._get_reportlab()
        Table = rl["Table"]
        TableStyle = rl["TableStyle"]
        mm = rl["mm"]

        data: list[list[Any]] = [[
            self._pdf_p("S.No", styles["th"]),
            self._pdf_p("Date", styles["th"]),
            self._pdf_p("Client", styles["th"]),
            self._pdf_p("Animal type", styles["th"]),
            self._pdf_p("Total Animals", styles["th"]),
            self._pdf_p("Unit Price service (per head) USD", styles["th"]),
            self._pdf_p("Total Service Revenue USD", styles["th"]),
            self._pdf_p("Unit Price offal USD", styles["th"]),
            self._pdf_p("Total Offal Revenue USD", styles["th"]),
            self._pdf_p("Total Combined Revenue USD", styles["th"]),
        ]]

        if rows:
            for index, item in enumerate(rows, start=1):
                data.append(
                    [
                        self._pdf_p(index, styles["td"]),
                        self._pdf_p(format_date(item.service_date), styles["td"]),
                        self._pdf_p(item.client_name or "", styles["td"]),
                        self._pdf_p(item.animal_type or "", styles["td"]),
                        self._pdf_p(format_int(item.total_animals), styles["td"]),
                        self._pdf_p(format_money_3(item.unit_price_per_head_usd), styles["td"]),
                        self._pdf_p(format_money_2(item.total_revenue_usd), styles["td"]),
                        self._pdf_p(format_money_3(item.unit_price_offal_usd), styles["td"]),
                        self._pdf_p(format_money_2(item.total_offal_revenue_usd), styles["td"]),
                        self._pdf_p(format_money_2(item.total_combined_revenue_usd), styles["td"]),
                    ]
                )
        else:
            data.append(
                [
                    self._pdf_p("No records found", styles["td"]),
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )

        service_row = len(data)
        data.append([
            self._pdf_p("Total Service Rev", styles["td_bold"]),
            "",
            "",
            "",
            "",
            "",
            self._pdf_p(format_money_2(totals.total_service_revenue_usd), styles["td_bold"]),
            "",
            "",
            "",
        ])

        offal_row = len(data)
        data.append([
            self._pdf_p("Total Offal Rev", styles["td_bold"]),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            self._pdf_p(format_money_2(totals.total_offal_revenue_usd), styles["td_bold"]),
            "",
        ])

        combined_row = len(data)
        data.append([
            self._pdf_p("Total Combined Rev", styles["td_bold"]),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            self._pdf_p(format_money_2(totals.total_combined_revenue_usd), styles["td_bold"]),
        ])

        table = Table(
            data,
            repeatRows=1,
            colWidths=[
                10 * mm,
                23 * mm,
                24 * mm,
                28 * mm,
                18 * mm,
                32 * mm,
                30 * mm,
                23 * mm,
                34 * mm,
                55 * mm,
            ],
            hAlign="LEFT",
        )

        style_commands = self._base_pdf_table_style()
        style_commands.extend([
            ("SPAN", (0, service_row), (5, service_row)),
            ("SPAN", (0, offal_row), (7, offal_row)),
            ("SPAN", (0, combined_row), (8, combined_row)),
        ])
        table.setStyle(TableStyle(style_commands))
        return table

    def _build_pdf_rows_table(
        self,
        rows: Sequence[SaaSReportRow],
        styles: dict[str, Any],
    ) -> Any:
        rl = self._get_reportlab()
        Table = rl["Table"]
        TableStyle = rl["TableStyle"]
        mm = rl["mm"]

        data: list[list[Any]] = [[
            self._pdf_p("S.No", styles["th"]),
            self._pdf_p("Date", styles["th"]),
            self._pdf_p("Client", styles["th"]),
            self._pdf_p("Animal type", styles["th"]),
            self._pdf_p("Animals", styles["th"]),
            self._pdf_p("Unit/head USD", styles["th"]),
            self._pdf_p("Service Rev USD", styles["th"]),
            self._pdf_p("Unit offal USD", styles["th"]),
            self._pdf_p("Offal Rev USD", styles["th"]),
            self._pdf_p("Combined Rev USD", styles["th"]),
        ]]

        if rows:
            for index, item in enumerate(rows, start=1):
                data.append(
                    [
                        self._pdf_p(index, styles["td"]),
                        self._pdf_p(format_date(item.service_date), styles["td"]),
                        self._pdf_p(item.client_name or "", styles["td"]),
                        self._pdf_p(item.animal_type or "", styles["td"]),
                        self._pdf_p(format_int(item.total_animals), styles["td"]),
                        self._pdf_p(format_money_3(item.unit_price_per_head_usd), styles["td"]),
                        self._pdf_p(format_money_2(item.total_revenue_usd), styles["td"]),
                        self._pdf_p(format_money_3(item.unit_price_offal_usd), styles["td"]),
                        self._pdf_p(format_money_2(item.total_offal_revenue_usd), styles["td"]),
                        self._pdf_p(format_money_2(item.total_combined_revenue_usd), styles["td"]),
                    ]
                )
        else:
            data.append([self._pdf_p("No records found", styles["td"])] + [""] * 9)

        table = Table(
            data,
            repeatRows=1,
            colWidths=[
                10 * mm,
                23 * mm,
                24 * mm,
                28 * mm,
                18 * mm,
                28 * mm,
                28 * mm,
                22 * mm,
                29 * mm,
                37 * mm,
            ],
            hAlign="LEFT",
        )
        table.setStyle(TableStyle(self._base_pdf_table_style()))
        return table

    def _build_pdf_totals_table(
        self,
        totals: SaaSReportTotals,
        styles: dict[str, Any],
    ) -> Any:
        rl = self._get_reportlab()
        Table = rl["Table"]
        TableStyle = rl["TableStyle"]
        mm = rl["mm"]

        data = [
            [self._pdf_p("Metric", styles["th"]), self._pdf_p("Value", styles["th"])],
            [self._pdf_p("Total Clients Served", styles["td"]), self._pdf_p(format_int(totals.total_clients_served), styles["td"])],
            [self._pdf_p("Total Animals", styles["td"]), self._pdf_p(format_int(totals.total_animals), styles["td"])],
            [self._pdf_p("Total Service Revenue USD", styles["td"]), self._pdf_p(format_money_2(totals.total_service_revenue_usd), styles["td"])],
            [self._pdf_p("Total Offal Revenue USD", styles["td"]), self._pdf_p(format_money_2(totals.total_offal_revenue_usd), styles["td"])],
            [self._pdf_p("Total Combined Revenue USD", styles["td"]), self._pdf_p(format_money_2(totals.total_combined_revenue_usd), styles["td"])],
        ]

        table = Table(
            data,
            colWidths=[70 * mm, 42 * mm],
            repeatRows=1,
            hAlign="LEFT",
        )
        table.setStyle(TableStyle(self._base_pdf_table_style()))
        return table

    def _build_pdf_client_summary_table(
        self,
        rows: Sequence[SaaSClientSummaryRow],
        styles: dict[str, Any],
    ) -> Any:
        rl = self._get_reportlab()
        Table = rl["Table"]
        TableStyle = rl["TableStyle"]
        mm = rl["mm"]

        data: list[list[Any]] = [[
            self._pdf_p("Client", styles["th"]),
            self._pdf_p("Rows", styles["th"]),
            self._pdf_p("Animals", styles["th"]),
            self._pdf_p("Service Rev", styles["th"]),
            self._pdf_p("Offal Rev", styles["th"]),
            self._pdf_p("Combined Rev", styles["th"]),
        ]]

        for item in rows:
            data.append(
                [
                    self._pdf_p(item.client_name or "", styles["td"]),
                    self._pdf_p(format_int(item.rows_count), styles["td"]),
                    self._pdf_p(format_int(item.total_animals), styles["td"]),
                    self._pdf_p(format_money_2(item.total_service_revenue_usd), styles["td"]),
                    self._pdf_p(format_money_2(item.total_offal_revenue_usd), styles["td"]),
                    self._pdf_p(format_money_2(item.total_combined_revenue_usd), styles["td"]),
                ]
            )

        table = Table(
            data,
            repeatRows=1,
            colWidths=[42 * mm, 18 * mm, 22 * mm, 28 * mm, 28 * mm, 30 * mm],
            hAlign="LEFT",
        )
        table.setStyle(TableStyle(self._base_pdf_table_style()))
        return table

    def _build_pdf_animal_summary_table(
        self,
        rows: Sequence[SaaSAnimalSummaryRow],
        styles: dict[str, Any],
    ) -> Any:
        rl = self._get_reportlab()
        Table = rl["Table"]
        TableStyle = rl["TableStyle"]
        mm = rl["mm"]

        data: list[list[Any]] = [[
            self._pdf_p("Animal Type", styles["th"]),
            self._pdf_p("Rows", styles["th"]),
            self._pdf_p("Animals", styles["th"]),
            self._pdf_p("Service Rev", styles["th"]),
            self._pdf_p("Offal Rev", styles["th"]),
            self._pdf_p("Combined Rev", styles["th"]),
        ]]

        for item in rows:
            data.append(
                [
                    self._pdf_p(item.animal_type or "", styles["td"]),
                    self._pdf_p(format_int(item.rows_count), styles["td"]),
                    self._pdf_p(format_int(item.total_animals), styles["td"]),
                    self._pdf_p(format_money_2(item.total_service_revenue_usd), styles["td"]),
                    self._pdf_p(format_money_2(item.total_offal_revenue_usd), styles["td"]),
                    self._pdf_p(format_money_2(item.total_combined_revenue_usd), styles["td"]),
                ]
            )

        table = Table(
            data,
            repeatRows=1,
            colWidths=[34 * mm, 18 * mm, 22 * mm, 28 * mm, 28 * mm, 30 * mm],
            hAlign="LEFT",
        )
        table.setStyle(TableStyle(self._base_pdf_table_style()))
        return table